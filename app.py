import hashlib
import streamlit as st
import io
import os
import re
import zipfile
from difflib import SequenceMatcher
import gspread
import openpyxl
from openpyxl.utils.cell import column_index_from_string
import json
import cv2
import numpy as np
import pandas as pd
import random
import base64
import time as pytime
import math
from google.oauth2 import service_account
from google.cloud import vision
from google.cloud import storage
from datetime import datetime, timedelta, timezone, time # ✅ เพิ่ม time
import string

# ✅ Roboflow Inference SDK สำหรับ Water Meter Detection
try:
    from inference_sdk import InferenceHTTPClient
    HAS_ROBOFLOW = True
except ImportError:
    HAS_ROBOFLOW = False
    InferenceHTTPClient = None

# ✅ 1.3: Daily Report Logging System
try:
    from daily_report_logger import (
        update_log_success,
        update_log_failed,
        print_daily_report,
        get_daily_summary,
    )
    HAS_LOGGER = True
except ImportError:
    HAS_LOGGER = False
    def update_log_success(*args, **kwargs): pass
    def update_log_failed(*args, **kwargs): pass
    def print_daily_report(*args, **kwargs): return "Logger not available"
    def get_daily_summary(*args, **kwargs): return {}

# =========================================================
# --- SQL SERVER IMPORTS (สำหรับ CUTEST SCADA Integration) ---
# =========================================================
try:
    import pyodbc
    HAS_PYODBC = True
except ImportError:
    HAS_PYODBC = False

try:
    import pymssql
    HAS_PYMSSQL = True
except ImportError:
    HAS_PYMSSQL = False
    
try:
    import sqlalchemy
    from sqlalchemy import create_engine, text
    HAS_SQLALCHEMY = True
except ImportError:
    HAS_SQLALCHEMY = False

# =========================
# Helpers / Utils (ต้องอยู่ก่อน UI)
# =========================
def make_thumb_data_url(image_bytes: bytes, max_size: int = 160, quality: int = 70) -> str:
    try:
        arr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return ""

        h, w = img.shape[:2]
        scale = min(max_size / max(h, w), 1.0)
        if scale < 1.0:
            img = cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)

        ok, buf = cv2.imencode(".jpg", img, [int(cv2.IMWRITE_JPEG_QUALITY), int(quality)])
        if not ok:
            return ""

        b64 = base64.b64encode(buf.tobytes()).decode("utf-8")
        return f"data:image/jpeg;base64,{b64}"
    except Exception:
        return ""
        
# =========================================================
# --- 📦 CONFIGURATION ---
# =========================================================
BUCKET_NAME = 'water-meter-images-watertreatmentplant'
FIXED_FOLDER_ID = '1XH4gKYb73titQLrgp4FYfLT2jzYRgUpO' 

# =========================================================
# --- 🕒 TIMEZONE HELPER ---
# =========================================================
def get_thai_time():
    """คืนค่าเวลาปัจจุบันในโซนไทย (UTC+7)"""
    tz = timezone(timedelta(hours=7))
    return datetime.now(tz)

# =========================================================
# --- PAGE CONFIG ---
# =========================================================
st.set_page_config(page_title="Smart Meter System", page_icon="💧", layout="centered")

st.markdown("""
    <style>
    .stButton>button { width: 100%; border-radius: 8px; font-weight: bold; }
    .status-box { padding: 15px; border-radius: 10px; margin-bottom: 15px; border: 1px solid #ddd; }
    .status-warning { background-color: #fff3cd; color: #856404; }
    .report-badge {
        background-color: #e3f2fd; color: #0d47a1;
        padding: 4px 8px; border-radius: 4px; font-size: 0.8em; font-weight: bold;
    }
    </style>
""", unsafe_allow_html=True)

# =========================================================
# --- CONFIGURATION & SECRETS ---
# =========================================================
if 'gcp_service_account' in st.secrets:
    try:
        key_dict = json.loads(st.secrets['gcp_service_account'])
        if 'private_key' in key_dict:
            key_dict['private_key'] = key_dict['private_key'].replace('\\n', '\n')

        creds = service_account.Credentials.from_service_account_info(
            key_dict,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
                "https://www.googleapis.com/auth/cloud-platform"
            ]
        )
    except Exception as e:
        st.error(f"❌ Error loading secrets: {e}")
        st.stop()
else:
    st.error("❌ Secrets not found.")
    st.stop()

gc = gspread.authorize(creds)
DB_SHEET_NAME = 'WaterMeter_System_DB'
REAL_REPORT_SHEET = 'FM-OP-01-10WaterReport'
VISION_CLIENT = vision.ImageAnnotatorClient(credentials=creds)
STORAGE_CLIENT = storage.Client(credentials=creds)

# =========================================================
# --- CLOUD STORAGE HELPERS ---
# =========================================================
def upload_image_to_storage(image_bytes, file_name):
    try:
        bucket = STORAGE_CLIENT.bucket(BUCKET_NAME)
        blob = bucket.blob(file_name)

        # ตั้งค่า Content-Type ให้ตรงกับนามสกุลไฟล์ (มือถือเปิดรูปได้ถูกต้อง)
        ext = str(file_name).lower().split(".")[-1] if "." in str(file_name) else "jpg"
        content_type = "image/png" if ext == "png" else "image/jpeg"

        blob.upload_from_string(image_bytes, content_type=content_type)
        return blob.public_url
    except Exception as e:
        return f"Error: {e}"


# =========================================================
# --- 🖼️ REFERENCE IMAGE (Auto Find) ---
# =========================================================
REF_IMAGE_FOLDER = "ref_images"

@st.cache_data(ttl=3600)
def load_ref_image_bytes_any(point_id: str):
    """
    หา reference รูปให้เอง:
    1) ref_images/POINT.(jpg/png)
    2) POINT.(jpg/png) ใน root
    3) ถ้าไม่เจอ → หาไฟล์ล่าสุดที่ขึ้นต้นด้วย POINT_ (เช่น POINT_2026...jpg)
    คืนค่า: (bytes, path) หรือ (None, None)
    """
    pid = str(point_id).strip().upper()
    bucket = STORAGE_CLIENT.bucket(BUCKET_NAME)

    # 1) ลองชื่อมาตรฐานก่อน
    candidates = []
    for ext in ["jpg", "jpeg", "png", "JPG", "JPEG", "PNG"]:
        candidates += [
            f"{REF_IMAGE_FOLDER}/{pid}.{ext}",
            f"{pid}.{ext}",
        ]

    # ลองดาวน์โหลดตรง ๆ (ไม่ต้องใช้ exists เพื่อกันเวอร์ชันไลบรารีต่างกัน)
    for path in candidates:
        try:
            blob = bucket.blob(path)
            data = blob.download_as_bytes()
            if data:
                return data, path
        except Exception:
            pass

    # 2) หาแบบ prefix เอาไฟล์ล่าสุด (POINT_....jpg/png)
    try:
        blobs = list(bucket.list_blobs(prefix=f"{pid}_"))
        blobs = [b for b in blobs if str(b.name).lower().endswith((".jpg", ".jpeg", ".png"))]
        if blobs:
            blobs.sort(key=lambda b: b.updated or datetime.min, reverse=True)
            b = blobs[0]
            data = b.download_as_bytes()
            return data, b.name
    except Exception:
        pass

    return None, None


# =========================================================
# --- SHEET HELPERS ---
# =========================================================
def col_to_index(col_str):
    col_str = str(col_str).upper().strip()
    num = 0
    for c in col_str:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

# ✅ แก้ไข: รับวันที่เข้ามาเพื่อหา Sheet เดือนที่ถูกต้อง (เผื่อลงย้อนหลังข้ามเดือน)
# ✅ Support both Thai and English month names with comprehensive month rollover handling
def get_thai_sheet_name(sh, target_date):
    """Find the correct monthly sheet based on target_date.
    
    Supports multiple naming conventions:
    - Thai months: ม.ค. 68, ก.พ. 68, etc.
    - English months: Jan2026, Feb2026, January2026, etc.
    
    Args:
        sh: gspread Spreadsheet object
        target_date: datetime.date object to find the correct month
    
    Returns:
        Sheet title if found, None otherwise
    """
    thai_months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    english_months_short = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    english_months_long = ["January", "February", "March", "April", "May", "June", 
                           "July", "August", "September", "October", "November", "December"]
    
    # ใช้วันที่ที่เลือก (target_date) แทนเวลาปัจจุบัน
    m_idx = target_date.month - 1
    # ปีพุทธศักราช (2 หรือ 4 หลัก)
    yy2 = str(target_date.year + 543)[-2:]
    yy4 = str(target_date.year + 543)
    # ปีค.ศ. (2 หรือ 4 หลัก)
    ad_yy2 = str(target_date.year)[-2:]
    ad_yy4 = str(target_date.year)
    
    all_sheets = [s.title for s in sh.worksheets()]
    
    # Try Thai month patterns first (highest priority)
    thai_patterns = [
        f"{thai_months[m_idx]}{yy2}",      # ม.ค.68
        f"{thai_months[m_idx][:-1]}{yy2}", # ม.ค68  (without dot)
        f"{thai_months[m_idx]} {yy2}",     # ม.ค. 68
        f"{thai_months[m_idx][:-1]} {yy2}",# ม.ค 68
        f"{thai_months[m_idx]}{yy4}",      # ม.ค.2568
        f"{thai_months[m_idx][:-1]}{yy4}", # ม.ค2568
        f"{thai_months[m_idx]} {yy4}",     # ม.ค. 2568
        f"{thai_months[m_idx][:-1]} {yy4}",# ม.ค 2568
    ]
    
    for p in thai_patterns:
        if p in all_sheets:
            return p
    
    # Try English month patterns (short names)
    eng_patterns = [
        f"{english_months_short[m_idx]}{yy2}",     # Jan69 (English + Thai Buddhist year)
        f"{english_months_short[m_idx]} {yy2}",    # Jan 69
        f"{english_months_short[m_idx]}{ad_yy4}",  # Jan2026
        f"{english_months_short[m_idx]} {ad_yy4}", # Jan 2026
        f"{english_months_short[m_idx]}{ad_yy2}",  # Jan26
        f"{english_months_short[m_idx]} {ad_yy2}", # Jan 26
        f"{english_months_long[m_idx]}{yy2}",      # January69
        f"{english_months_long[m_idx]} {yy2}",     # January 69
        f"{english_months_long[m_idx]}{ad_yy4}",   # January2026
        f"{english_months_long[m_idx]} {ad_yy4}",  # January 2026
        f"{english_months_long[m_idx]}{ad_yy2}",   # January26
        f"{english_months_long[m_idx]} {ad_yy2}",  # January 26
    ]
    
    for p in eng_patterns:
        if p in all_sheets:
            return p
    
    return None

def find_day_row_exact(ws, day: int):
    col = ws.col_values(1)
    for i, v in enumerate(col, start=1):
        try:
            if int(str(v).strip()) == int(day):
                return i
        except:
            pass
    return None

@st.cache_data(ttl=300)
def load_points_master():
    sh = gc.open(DB_SHEET_NAME)
    ws = sh.worksheet("PointsMaster")
    return ws.get_all_records()

def safe_int(x, default=0):
    try: return int(float(x)) if x and str(x).strip() else default
    except: return default

def safe_float(x, default=0.0):
    try: return float(x) if x and str(x).strip() else default
    except: return default

def parse_bool(v):
    if v is None: return False
    return str(v).strip().lower() in ("true", "1", "yes", "y", "t", "on")

def get_meter_config(point_id):
    try:
        records = load_points_master()
        pid = str(point_id).strip().upper()
        for item in records:
            if str(item.get('point_id', '')).strip().upper() == pid:
                item['decimals'] = safe_int(item.get('decimals'), 0)
                item['keyword'] = str(item.get('keyword', '')).strip()
                exp = safe_int(item.get('expected_digits'), 0)
                if exp == 0: exp = safe_int(item.get('int_digits'), 0)
                item['expected_digits'] = exp
                item['report_col'] = str(item.get('report_col', '')).strip()
                item['ignore_red'] = parse_bool(item.get('ignore_red'))
                item['roi_x1'] = safe_float(item.get('roi_x1'), 0.0)
                item['roi_y1'] = safe_float(item.get('roi_y1'), 0.0)
                item['roi_x2'] = safe_float(item.get('roi_x2'), 0.0)
                item['roi_y2'] = safe_float(item.get('roi_y2'), 0.0)
                item['type'] = str(item.get('type', '')).strip()
                item['name'] = str(item.get('name', '')).strip()
                return item
        return None
    except: return None

# ✅ แก้ไข: รับ target_date เพื่อลงให้ถูกวัน
def export_to_real_report(point_id, read_value, inspector, report_col, target_date, debug=False):
    """ส่งค่าลง Google Sheet REAL_REPORT_SHEET
    - debug=False: คืน True/False เหมือนเดิม
    - debug=True : คืน (ok, message) เพื่อโชว์สาเหตุเวลาส่งไม่เข้า
    """

    def _ret(ok, msg=""):
        return (ok, msg) if debug else ok

    if not report_col:
        return _ret(False, "report_col ว่าง")
    report_col = str(report_col).strip()
    if report_col in ("-", "—", "–"):
        return _ret(False, "report_col เป็น '-' (ยังไม่ได้ตั้งค่าใน PointsMaster)")

    # เปิดชีท
    try:
        sh = gc.open(REAL_REPORT_SHEET)
    except Exception as e:
        return _ret(False, f"เปิดชีท '{REAL_REPORT_SHEET}' ไม่ได้: {e}")

    # หาแท็บเดือน
    sheet_name = None
    try:
        sheet_name = get_thai_sheet_name(sh, target_date)
    except Exception:
        sheet_name = None

    # ถ้าไม่เจอ → หาแบบฟัซซี่ (ตัดช่องว่าง/จุด)
    if not sheet_name:
        try:
            thai_months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
            english_months_short = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            english_months_long = ["January", "February", "March", "April", "May", "June", 
                                   "July", "August", "September", "October", "November", "December"]
            
            m_idx = target_date.month - 1
            yy2_thai = str(target_date.year + 543)[-2:]
            yy4_thai = str(target_date.year + 543)
            yy2_ad = str(target_date.year)[-2:]
            yy4_ad = str(target_date.year)
            
            m_norm_thai = thai_months[m_idx].replace(".", "").replace(" ", "")

            def norm(x):
                return str(x).replace(".", "").replace(" ", "").lower().strip()

            for t in [s.title for s in sh.worksheets()]:
                tn = norm(t)
                # Check Thai month patterns
                if (m_norm_thai.lower() in tn) and (yy2_thai in tn or yy4_thai in tn):
                    sheet_name = t
                    break
                # Check English month patterns (case-insensitive)
                if (english_months_short[m_idx].lower() in tn or english_months_long[m_idx].lower() in tn) and (yy2_ad in tn or yy4_ad in tn):
                    sheet_name = t
                    break
        except Exception:
            sheet_name = None

    # เปิด worksheet
    try:
        ws = sh.worksheet(sheet_name) if sheet_name else sh.get_worksheet(0)
    except Exception as e:
        return _ret(False, f"เปิดแท็บ '{sheet_name}' ไม่ได้: {e}")

    # หาแถวของวัน
    try:
        target_day = int(target_date.day)
        target_row = find_day_row_exact(ws, target_day) or (6 + target_day)
    except Exception as e:
        return _ret(False, f"หาแถวของวันไม่สำเร็จ: {e}")

    # หา col
    target_col = col_to_index(report_col)
    if target_col == 0:
        return _ret(False, f"report_col '{report_col}' แปลงเป็นคอลัมน์ไม่ได้")

    # เขียนค่า
    try:
        ws.update_cell(target_row, target_col, read_value)
        return _ret(True, f"OK → sheet='{ws.title}', row={target_row}, col={report_col}({target_col}), val={read_value}")
    except Exception as e:
        return _ret(False, f"เขียนค่าไม่สำเร็จ: {e}")



# ✅ แก้ไข: รับ target_date เพื่อลง Timestamp ให้ถูกวัน

# =========================================================
# --- 🚀 QUOTA-SAFE BATCH HELPERS (Sheets) ---
# =========================================================
def _is_quota_429(err: Exception) -> bool:
    msg = str(err)
    return ("429" in msg) or ("Quota exceeded" in msg) or ("Read requests" in msg)

def _with_retry(fn, *args, max_retries: int = 6, base_sleep: float = 0.8, **kwargs):
    """
    Retry wrapper for Google Sheets calls that may hit 429 quota.
    - Exponential backoff + jitter
    """
    last_err = None
    for i in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            last_err = e
            if _is_quota_429(e) and i < max_retries - 1:
                # backoff: 0.8, 1.6, 3.2, ...
                sleep_s = base_sleep * (2 ** i) + random.random() * 0.4
                pytime.sleep(sleep_s)
                continue
            raise
    if last_err:
        raise last_err

def export_many_to_real_report_batch(items: list, target_date, debug: bool = False, write_mode: str = "overwrite"):
    """
    Export หลายจุดลง WaterReport ด้วย 1 batch_update (ลด Read/Write requests มาก ๆ)
    items: list[dict] ต้องมี keys: point_id, value, report_col
    คืนค่า:
      - ok_pids: list[str]
      - fail_list: list[tuple(pid, reason)]
    """
    ok_pids = []
    fail_list = []

    if not items:
        return ok_pids, fail_list

    # เปิดชีทครั้งเดียว + retry กัน quota
    try:
        sh = _with_retry(gc.open, REAL_REPORT_SHEET)
    except Exception as e:
        reason = f"เปิดชีท '{REAL_REPORT_SHEET}' ไม่ได้: {e}"
        for it in items:
            fail_list.append((it.get("point_id", ""), reason))
        return ok_pids, fail_list

    # หาแท็บเดือนครั้งเดียว
    sheet_name = None
    try:
        sheet_name = get_thai_sheet_name(sh, target_date)
    except Exception:
        sheet_name = None

    # fallback แบบฟัซซี่ (ครั้งเดียว)
    if not sheet_name:
        try:
            thai_months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
            m_idx = target_date.month - 1
            yy2 = str(target_date.year + 543)[-2:]
            yy4 = str(target_date.year + 543)
            m_norm = thai_months[m_idx].replace(".", "").replace(" ", "")

            def norm(x):
                return str(x).replace(".", "").replace(" ", "").strip()

            for t in [s.title for s in sh.worksheets()]:
                tn = norm(t)
                if (m_norm in tn) and (yy2 in tn or yy4 in tn):
                    sheet_name = t
                    break
        except Exception:
            sheet_name = None

    # เปิด worksheet ครั้งเดียว + retry
    try:
        ws = _with_retry(sh.worksheet, sheet_name) if sheet_name else _with_retry(sh.get_worksheet, 0)
    except Exception as e:
        reason = f"เปิดแท็บ '{sheet_name}' ไม่ได้: {e}"
        for it in items:
            fail_list.append((it.get("point_id", ""), reason))
        return ok_pids, fail_list

    # ✅ ลด Read requests: ใช้สูตร row แบบคงที่ (ถ้า template เปลี่ยนค่อยเปิด find_day_row_exact อีกที)
    try:
        target_row = 6 + int(target_date.day)
    except Exception:
        target_row = 6 + 1

    # ---- กันเขียนทับ: ถ้าเลือก 'เขียนเฉพาะช่องว่าง' จะอ่านค่าเดิมในแถวนี้ก่อน 1 ครั้ง ----
    existing_row = None
    wm = str(write_mode or 'overwrite').strip().lower()
    if wm in ('empty_only', 'skip_non_empty', 'no_overwrite', 'nooverwrite', 'blank_only'):
        try:
            existing_row = _with_retry(ws.row_values, target_row)
        except Exception:
            existing_row = None

    # เตรียม batch ranges
    data = []
    for it in items:
        pid = str(it.get("point_id", "")).strip().upper()
        report_col = str(it.get("report_col", "")).strip()
        val = it.get("value", "")

        if not report_col or report_col in ("-", "—", "–"):
            fail_list.append((pid, "report_col ว่าง/เป็น '-' ใน PointsMaster"))
            continue

        target_col = col_to_index(report_col)
        if target_col <= 0:
            fail_list.append((pid, f"report_col '{report_col}' แปลงคอลัมน์ไม่ได้"))
            continue

        # ถ้าเลือก 'เขียนเฉพาะช่องว่าง' และช่องมีข้อมูลแล้ว -> ข้าม
        if existing_row is not None:
            try:
                existing_val = existing_row[target_col - 1] if (target_col - 1) < len(existing_row) else ''
                if str(existing_val).strip() != '':
                    fail_list.append((pid, 'SKIP_NON_EMPTY'))
                    continue
            except Exception:
                pass

        # A1 เช่น "Y18"
        a1 = gspread.utils.rowcol_to_a1(target_row, target_col)
        data.append({"range": a1, "values": [[val]]})
        ok_pids.append(pid)

    if not data:
        return [], fail_list

    # batch_update ครั้งเดียว + retry กัน quota
    try:
        _with_retry(ws.batch_update, data, value_input_option="USER_ENTERED")
        return ok_pids, fail_list
    except Exception as e:
        # ถ้า batch fail → ถือว่าทั้งหมด fail (ให้ user กดใหม่ได้)
        reason = f"เขียนค่าไม่สำเร็จ (batch_update): {e}"
        for pid in ok_pids:
            fail_list.append((pid, reason))
        return [], fail_list

def append_rows_dailyreadings_batch(rows: list):
    """
    append_rows ลง DailyReadings ครั้งเดียว (ลด requests)
    rows: list[list] แต่ละแถวต้องตรงกับ schema DailyReadings
    คืนค่า (ok:bool, message:str)
    """
    if not rows:
        return True, "NO_ROWS"

    try:
        sh = _with_retry(gc.open, DB_SHEET_NAME)
        ws = _with_retry(sh.worksheet, "DailyReadings")
        _with_retry(ws.append_rows, rows, value_input_option="USER_ENTERED")
        return True, f"APPENDED {len(rows)}"
    except Exception as e:
        return False, str(e)
        
# =========================================================
# --- ✅ WATERREPORT PROGRESS (92 จุด) ---
# =========================================================
@st.cache_data(ttl=60)
def get_waterreport_progress_snapshot(target_date):
    """
    เช็คความคืบหน้าการลงค่าใน REAL_REPORT_SHEET ของ 'วันนั้น'
    - total = จำนวนจุดที่ "ตั้ง report_col แล้ว" (เช็คได้จริงใน WaterReport)
    - total_all = จำนวน point_id ทั้งหมดใน PointsMaster
    """
    pm = load_points_master() or []

    expected_all = []
    expected_report = []
    missing_config = []
    seen = set()

    for it in pm:
        pid = str(it.get("point_id", "")).strip().upper()
        if not pid or pid in seen:
            continue
        seen.add(pid)

        report_col = str(it.get("report_col", "")).strip()
        name = str(it.get("name", "") or "").strip()
        rec = {"point_id": pid, "report_col": report_col, "name": name}

        expected_all.append(rec)

        if report_col and report_col not in ("-", "—", "–"):
            # กัน report_col แปลเป็นคอลัมน์ไม่ได้
            if col_to_index(report_col) > 0:
                expected_report.append(rec)
            else:
                missing_config.append({**rec, "reason": "BAD_REPORT_COL"})
        else:
            missing_config.append({**rec, "reason": "NO_REPORT_COL"})

    total_all = len(expected_all)
    total_report = len(expected_report)
    asof = get_thai_time().strftime("%Y-%m-%d %H:%M:%S")

    # --- open spreadsheet ---
    try:
        sh = _with_retry(gc.open, REAL_REPORT_SHEET)
    except Exception as e:
        return {
            "ok": False,
            "total": total_report,
            "total_report": total_report,
            "total_all": total_all,
            "config_missing": len(missing_config),
            "filled": 0,
            "missing": expected_report,
            "done_set": set(),
            "value_map": {},
            "sheet_title": None,
            "row": None,
            "asof": asof,
            "error": f"open REAL_REPORT_SHEET failed: {e}",
        }

    # --- find month sheet ---
    sheet_name = None
    try:
        sheet_name = get_thai_sheet_name(sh, target_date)
    except Exception:
        sheet_name = None

    try:
        ws = _with_retry(sh.worksheet, sheet_name) if sheet_name else _with_retry(sh.get_worksheet, 0)
    except Exception as e:
        return {
            "ok": False,
            "total": total_report,
            "total_report": total_report,
            "total_all": total_all,
            "config_missing": len(missing_config),
            "filled": 0,
            "missing": expected_report,
            "done_set": set(),
            "value_map": {},
            "sheet_title": sheet_name,
            "row": None,
            "asof": asof,
            "error": f"open worksheet failed: {e}",
        }

    # --- read row of day ---
    try:
        target_row = 6 + int(target_date.day)
    except Exception:
        target_row = 7

    try:
        row_vals = _with_retry(ws.row_values, target_row)
    except Exception as e:
        return {
            "ok": False,
            "total": total_report,
            "total_report": total_report,
            "total_all": total_all,
            "config_missing": len(missing_config),
            "filled": 0,
            "missing": expected_report,
            "done_set": set(),
            "value_map": {},
            "sheet_title": ws.title,
            "row": target_row,
            "asof": asof,
            "error": f"read row_values failed: {e}",
        }

    done_set = set()
    value_map = {}
    missing = []

    for it in expected_report:
        pid = it["point_id"]
        col_idx = col_to_index(it["report_col"])
        if col_idx <= 0:
            missing.append({**it, "reason": "BAD_REPORT_COL"})
            continue

        existing = row_vals[col_idx - 1] if (col_idx - 1) < len(row_vals) else ""
        if str(existing).strip() != "":
            done_set.add(pid)
            value_map[pid] = existing
        else:
            missing.append(it)

    filled = len(done_set)

    return {
        "ok": True,
        "total": total_report,
        "total_report": total_report,
        "total_all": total_all,
        "config_missing": len(missing_config),
        "filled": filled,
        "missing": missing,
        "done_set": done_set,
        "value_map": value_map,
        "sheet_title": ws.title,
        "row": target_row,
        "asof": asof,
        "error": "",
    }

# =========================================================
# --- SQL SERVER INTEGRATION (CUTEST SCADA 2018) ---
# =========================================================
def test_sql_connection(server: str, database: str, username: str, password: str) -> tuple[bool, str]:
    """
    ทดสอบเชื่อมต่อ SQL Server (รองรับ pyodbc, pymssql, sqlalchemy)
    คืนค่า: (success: bool, message: str)
    """
    # 🔹 Attempt 1: pymssql (แนะนำสำหรับ macOS/Linux)
    if HAS_PYMSSQL:
        try:
            conn = pymssql.connect(
                host=server,
                user=username,
                password=password,
                database=database,
                timeout=5
            )
            cursor = conn.cursor()
            cursor.execute("SELECT @@version")
            result = cursor.fetchone()
            conn.close()
            version = result[0] if result else "Unknown"
            return True, f"✅ เชื่อมต่อสำเร็จ (pymssql)\n{str(version)[:100]}"
        except Exception as e:
            pass
    
    # 🔹 Attempt 2: sqlalchemy
    if HAS_SQLALCHEMY:
        try:
            conn_str = f"mssql+pyodbc://{username}:{password}@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server"
            engine = create_engine(conn_str, pool_pre_ping=True)
            with engine.connect() as conn:
                result = conn.execute(text("SELECT @@version"))
                version = result.fetchone()[0]
                return True, f"✅ เชื่อมต่อสำเร็จ (sqlalchemy)\n{str(version)[:100]}"
        except Exception as e:
            pass
    
    # 🔹 Attempt 3: pyodbc (Windows)
    if HAS_PYODBC:
        try:
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password};Connection Timeout=5"
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()
            cursor.execute("SELECT @@version")
            result = cursor.fetchone()
            conn.close()
            return True, f"✅ เชื่อมต่อสำเร็จ (pyodbc)\n{result[0][:100]}"
        except Exception as e:
            pass
    
    return False, "❌ ไม่มี driver SQL ที่ติดตั้ง\nต้องติดตั้ง: pip install pymssql"

def query_scada_values(
    server: str, 
    database: str, 
    username: str, 
    password: str,
    point_id: str,
    target_date: datetime.date,
    target_time: str = None
) -> dict:
    """
    ดึงค่าจาก CUTEST SCADA SQL Server
    รองรับ: pymssql, sqlalchemy, pyodbc
    
    คืนค่า: {
        "success": bool,
        "value": float or None,
        "timestamp": str,
        "message": str,
        "all_records": list
    }
    """
    date_str = target_date.strftime("%Y-%m-%d")
    table_candidates = ["History_Data", "Readings", "dbo.History_Data", "dbo.Readings", "TagHistory"]
    
    # 🔹 Method 1: pymssql
    if HAS_PYMSSQL:
        try:
            conn = pymssql.connect(
                host=server,
                user=username,
                password=password,
                database=database,
                timeout=10
            )
            cursor = conn.cursor()
            
            for table in table_candidates:
                try:
                    query = f"""
                    SELECT TOP 100 TagName, Value, Timestamp 
                    FROM {table}
                    WHERE TagName LIKE '%{point_id}%'
                      AND CAST(Timestamp AS DATE) = '{date_str}'
                    ORDER BY Timestamp DESC
                    """
                    cursor.execute(query)
                    query_result = cursor.fetchall()
                    if query_result:
                        conn.close()
                        return _process_sql_results(query_result, table, date_str, point_id)
                except:
                    continue
            
            conn.close()
            return {
                "success": False,
                "message": f"ไม่พบข้อมูล point_id='{point_id}' ในวันที่ {date_str}",
                "all_records": []
            }
        except Exception as e:
            pass
    
    # 🔹 Method 2: sqlalchemy
    if HAS_SQLALCHEMY:
        try:
            conn_str = f"mssql+pyodbc://{username}:{password}@{server}/{database}?driver=ODBC+Driver+17+for+SQL+Server"
            engine = create_engine(conn_str, pool_pre_ping=True)
            
            for table in table_candidates:
                try:
                    query = f"""
                    SELECT TOP 100 TagName, Value, Timestamp 
                    FROM {table}
                    WHERE TagName LIKE '%{point_id}%'
                      AND CAST(Timestamp AS DATE) = '{date_str}'
                    ORDER BY Timestamp DESC
                    """
                    with engine.connect() as conn:
                        result = conn.execute(text(query))
                        rows = result.fetchall()
                        if rows:
                            query_result = [(row[0], row[1], row[2]) for row in rows]
                            return _process_sql_results(query_result, table, date_str, point_id)
                except:
                    continue
            
            return {
                "success": False,
                "message": f"ไม่พบข้อมูล point_id='{point_id}' ในวันที่ {date_str}",
                "all_records": []
            }
        except Exception as e:
            pass
    
    # 🔹 Method 3: pyodbc
    if HAS_PYODBC:
        try:
            conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"
            conn = pyodbc.connect(conn_str, timeout=10)
            cursor = conn.cursor()
            
            for table in table_candidates:
                try:
                    query = f"""
                    SELECT TOP 100 TagName, Value, Timestamp 
                    FROM {table}
                    WHERE TagName LIKE '%{point_id}%'
                      AND CAST(Timestamp AS DATE) = '{date_str}'
                    ORDER BY Timestamp DESC
                    """
                    cursor.execute(query)
                    query_result = cursor.fetchall()
                    if query_result:
                        conn.close()
                        return _process_sql_results(query_result, table, date_str, point_id)
                except:
                    continue
            
            conn.close()
            return {
                "success": False,
                "message": f"ไม่พบข้อมูล point_id='{point_id}' ในวันที่ {date_str}",
                "all_records": []
            }
        except Exception as e:
            pass
    
    return {
        "success": False,
        "message": "❌ ไม่มี SQL driver ที่ติดตั้ง\nต้องติดตั้ง: pip install pymssql",
        "all_records": []
    }

def _process_sql_results(query_result, table_found, date_str, point_id):
    """ตัวช่วยสำหรับประมวลผลผลลัพธ์ SQL"""
    records = []
    latest_value = None
    latest_time = None
    
    for row in query_result:
        tag_name = row[0]
        value = row[1]
        timestamp = row[2]
        records.append({
            "tag": tag_name,
            "value": value,
            "timestamp": str(timestamp)
        })
        if latest_value is None:
            latest_value = value
            latest_time = str(timestamp)
    
    return {
        "success": True,
        "value": latest_value,
        "timestamp": latest_time,
        "table": table_found,
        "record_count": len(records),
        "all_records": records,
        "message": f"✅ พบข้อมูล {len(records)} แถว จากตาราง {table_found}"
    }

def save_to_db(point_id, inspector, meter_type, manual_val, ai_val, status, target_date, image_url="-"):
    try:
        sh = gc.open(DB_SHEET_NAME)
        ws = sh.worksheet("DailyReadings")
        
        # สร้าง Timestamp: วันที่เลือก + เวลาปัจจุบัน (เพื่อให้รู้ว่าคีย์ตอนกี่โมง แต่วันที่เป็นของวันที่เลือก)
        current_time = get_thai_time().time()
        record_timestamp = datetime.combine(target_date, current_time)
        
        row = [record_timestamp.strftime("%Y-%m-%d %H:%M:%S"), meter_type, point_id, inspector, manual_val, ai_val, status, image_url]
        ws.append_row(row)
        return True
    except: return False

# =========================================================
# --- 🧠 OCR ENGINE (Clean & Robust) ---
# =========================================================

# ------------------------ SCADA Excel Upload (Export) ------------------------
def _normalize_scada_time(value):
    """
    แปลงเวลาให้เป็นรูปแบบ 'HH:MM' เพื่อเทียบกันง่าย (รองรับ time/datetime/str/float)
    
    ✅ 24:00 standardization:
    - 24:00 → 23:55 (standard for end-of-day)
    """
    import datetime as _dt
    if value is None:
        return None

    # Excel time (เช่น 0.9965) = สัดส่วนของวัน
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        seconds = int(round(float(value) * 24 * 60 * 60))
        h = (seconds // 3600) % 24
        m = (seconds % 3600) // 60
        result = f"{h:02d}:{m:02d}"
        # Apply 24:00 → 23:55 conversion
        if h == 24 and m == 0:
            return "23:55"
        return result

    if isinstance(value, _dt.datetime):
        value = value.time()
    if isinstance(value, _dt.time):
        result = f"{value.hour:02d}:{value.minute:02d}"
        # Apply 24:00 → 23:55 conversion
        if value.hour == 24 and value.minute == 0:
            return "23:55"
        return result

    s = str(value).strip()
    # 23.55 or 24.00
    if re.match(r"^\d{1,2}\.\d{2}$", s):
        h, m = s.split(".")
        h = int(h)
        m = int(m)
        # ✅ 24:00 → 23:55 conversion
        if h == 24 and m == 0:
            return "23:55"
        return f"{h:02d}:{m:02d}"
    
    # 23:55 or 24:00 or 23:55:00 or 24:00:00
    if re.match(r"^\d{1,2}:\d{2}", s):
        parts = s.split(":")
        h = int(parts[0])
        m = int(parts[1])
        # ✅ 24:00 → 23:55 conversion
        if h == 24 and m == 0:
            return "23:55"
        return f"{h:02d}:{m:02d}"

    return None


def _strip_date_prefix(name: str) -> str:
    """
    เอาวันที่นำหน้าออก (เช่น 2026_01_12_Daily_Report -> Daily_Report)
    """
    base = os.path.splitext(os.path.basename(name))[0]
    base = re.sub(r"^\d{4}_\d{2}_\d{2}_", "", base)
    return base.strip().lower()


def load_scada_excel_mapping(local_path: str = "DB_Water_Scada.xlsx", uploaded_bytes=None):
    """
    อ่าน mapping จากไฟล์ DB_Water_Scada.xlsx
    ต้องมีหัวตาราง: PointID, File, Sheet, Time, Colume
    คืนค่าเป็น list ของ dict: {point_id, file_key, sheet, time, col}
    """
    if uploaded_bytes:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_bytes), data_only=True)
    else:
        if not os.path.exists(local_path):
            return []
        wb = openpyxl.load_workbook(local_path, data_only=True)

    ws = wb[wb.sheetnames[0]]

    # หาแถวหัวตาราง
    header_row = None
    header_map = {}
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        row_str = [str(v).strip().lower() if v is not None else "" for v in row_vals]
        if "pointid" in row_str and "file" in row_str and "sheet" in row_str:
            header_row = r
            for idx, name in enumerate(row_str, start=1):
                if name in ["pointid", "file", "sheet", "time", "colume", "column"]:
                    header_map[name] = idx
            break

    if not header_row:
        return []

    # รองรับสะกด Colume/Column
    col_idx = header_map.get("colume") or header_map.get("column")
    out = []
    for r in range(header_row + 1, ws.max_row + 1):
        point_id = ws.cell(r, header_map["pointid"]).value
        if point_id is None or str(point_id).strip() == "":
            continue

        file_key = ws.cell(r, header_map["file"]).value
        sheet = ws.cell(r, header_map["sheet"]).value
        t = ws.cell(r, header_map.get("time", 0)).value if header_map.get("time") else None
        col = ws.cell(r, col_idx).value if col_idx else None

        out.append({
            "point_id": str(point_id).strip(),
            "file_key": str(file_key).strip() if file_key is not None else "",
            "sheet": str(sheet).strip() if sheet is not None else "Sheet1",
            "time": t,
            "col": str(col).strip() if col is not None else "",
        })
    return out


def _find_cell_exact(ws, target_text: str, max_rows=60, max_cols=40):
    target = target_text.strip().lower()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == target:
                return r, c
    return None


def _hhmm_to_minutes(hhmm: str, normalize_24_00=True):
    """
    แปลง HH:MM เป็นจำนวนนาทีตั้งแต่เที่ยงคืน
    
    ✅ 24:00 standardization:
    - 24:00 ถือว่าเป็น 23:55 (สุดท้ายของวัน)
    - กำหนดมาตรฐาน: "24:00 ของวัน D" = "23:55 ของวัน D"
    
    Args:
        hhmm: string format "HH:MM" (e.g., "24:00", "23:55")
        normalize_24_00: if True, convert 24:00 → 23:55
    
    Returns:
        minutes since midnight, or None if invalid
    """
    try:
        h, m = str(hhmm).split(":")
        h = int(h)
        m = int(m)
        
        # ✅ Handle 24:00 normalization
        if normalize_24_00 and h == 24 and m == 0:
            # 24:00 ของวัน D = 23:55 ของวัน D
            return 23 * 60 + 55
        
        # Validate time range
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        
        return h * 60 + m
    except Exception:
        return None


def _minutes_to_hhmm(minutes: int) -> str:
    """
    แปลงนาทีมาเป็น HH:MM format
    """
    try:
        h = minutes // 60
        m = minutes % 60
        return f"{h:02d}:{m:02d}"
    except Exception:
        return None


def _normalize_time_to_standard(hhmm: str) -> str:
    """
    Normalize any time format to standard HH:MM
    
    ✅ 24:00 standardization (สำคัญ):
    - Input: "24:00" → Output: "23:55"
    - This is the company standard for end-of-day
    
    Returns:
        Normalized time string, or None if invalid
    """
    try:
        # First normalize to HH:MM
        normalized = _normalize_scada_time(hhmm)
        if not normalized:
            return None
        
        h, m = normalized.split(":")
        h = int(h)
        m = int(m)
        
        # ✅ Apply 24:00 → 23:55 conversion
        if h == 24 and m == 0:
            return "23:55"
        
        if h < 0 or h > 23 or m < 0 or m > 59:
            return None
        
        return f"{h:02d}:{m:02d}"
    except Exception:
        return None


def _find_nearest_time_row(time_rows: list, target_minutes: int, max_diff_minutes: int = 300) -> int:
    """
    Find the row with time closest to target_minutes (nearest time algorithm)
    
    ✅ Key feature: Handles missing data by finding nearest available time
    
    Args:
        time_rows: list of tuples (row_number, minutes_since_midnight)
        target_minutes: target time in minutes (e.g., 1435 for 23:55)
        max_diff_minutes: max allowed difference (default 5 mins = 300 sec)
    
    Returns:
        row_number if found, None otherwise
    
    Example:
        time_rows = [(10, 1430), (11, 1435), (12, 1440)]  # 23:50, 23:55, 24:00→23:55
        target = 1435  # 23:55
        → returns 11 (exact match)
        
        If 23:55 data missing, still finds nearest (23:50 or 00:00)
    """
    if not time_rows:
        return None
    
    if target_minutes is None:
        # No target specified, return last available
        return time_rows[-1][0]
    
    # Find closest match
    nearest = min(time_rows, key=lambda x: abs(x[1] - target_minutes))
    diff = abs(nearest[1] - target_minutes)
    
    # Only return if within acceptable range
    if diff <= max_diff_minutes:
        return nearest[0]
    
    # If no match within range, return last available (fallback)
    return time_rows[-1][0]



def _extract_value_from_ws(ws, target_time_hhmm, value_col_letter: str, time_header="Time", max_scan_rows: int = 5000):
    """
    ดึงค่าจากตารางที่มีคอลัมน์เวลา (Time) โดย:
    - หา header 'Time' ก่อน
    - สแกนแถวข้อมูลจำนวนจำกัด (กันไฟล์ใหญ่ max_row หลอก)
    - เลือกแถวที่ใกล้เวลาเป้าหมายที่สุด (หรือแถวสุดท้าย)
    - ถ้า cell ว่าง ไล่ขึ้นไปหาแถวก่อนหน้าที่มีค่า
    คืนค่า: (value, status)
    """
    hdr = _find_cell_exact(ws, time_header)
    if not hdr:
        return None, "NO_TIME_HEADER"

    hdr_row, time_col = hdr

    # เก็บแถวที่มีเวลา (จำกัดจำนวนแถวที่สแกน)
    time_rows = []
    blank_streak = 0
    max_r = min(ws.max_row or 0, hdr_row + max_scan_rows)
    for r in range(hdr_row + 1, max_r + 1):
        v = ws.cell(r, time_col).value
        hhmm = _normalize_scada_time(v)
        mm = _hhmm_to_minutes(hhmm) if hhmm else None

        if mm is not None:
            time_rows.append((r, mm))
            blank_streak = 0
        else:
            blank_streak += 1
            # ถ้าเริ่มเจอแถวว่างยาว ๆ และมีข้อมูลแล้ว ให้หยุด เพื่อความเร็ว
            if blank_streak >= 80 and time_rows:
                break

    if not time_rows:
        return None, "NO_DATA_ROW"

    # เลือกแถวที่ “ใกล้เวลาเป้าหมายที่สุด”
    if target_time_hhmm:
        tmm = _hhmm_to_minutes(target_time_hhmm)
        target_row = _find_nearest_time_row(time_rows, tmm, max_diff_minutes=300)
        if target_row is None:
            target_row = time_rows[-1][0]
    else:
        target_row = time_rows[-1][0]

    # คอลัมน์ค่า
    try:
        col_idx = column_index_from_string(str(value_col_letter).strip().upper())
    except Exception:
        return None, "BAD_COLUMN"

    # ถ้าแถวที่เลือกว่าง → ไล่ขึ้นไปหาแถวก่อนหน้าที่มีค่า
    for rr in range(target_row, hdr_row, -1):
        val = ws.cell(rr, col_idx).value
        if val not in (None, "", " "):
            return val, "OK"

    return None, "EMPTY_CELL"


def _norm_filekey(name: str) -> str:
    """normalize ชื่อไฟล์/คีย์เพื่อเทียบกันแบบหยาบ ๆ"""
    base = os.path.splitext(os.path.basename(str(name)))[0]
    base = base.strip().lower()
    base = re.sub(r"\s+", "_", base)
    base = re.sub(r"[^a-z0-9_]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")
    return base

def _is_uf_gen_report_workbook(wb) -> bool:
    """ตรวจว่าเป็นไฟล์ UF/System แบบใหม่ (เช่น AF_Report_Gen.. มีหลาย sheet: Total/PV/FM_01..)"""
    try:
        names = {str(n).strip().lower() for n in (wb.sheetnames or [])}
        return ("total" in names) and ("pv" in names) and any(n.startswith("fm_") for n in names)
    except Exception:
        return False

def _resolve_sheet_name_for_export(wb, desired_sheet: str, point_id: str) -> str:
    """
    map ชื่อ sheet ให้เข้ากับไฟล์จริง:
    - ถ้ามี sheet ตรงชื่อ -> ใช้เลย
    - ถ้า desired='Sheet1' แต่ไฟล์เป็น UF gen report -> ใช้ 'Total' (เทียบเท่า Sheet1 เดิม)
    - ไม่งั้น fallback เป็น sheet แรก
    """
    try:
        if not wb:
            return desired_sheet
        sheetnames = wb.sheetnames or []
        if desired_sheet in sheetnames:
            return desired_sheet

        # case-insensitive match
        ds = str(desired_sheet or "").strip().lower()
        for s in sheetnames:
            if str(s).strip().lower() == ds:
                return s

        # UF gen report: Sheet1 -> Total
        if ds in ("sheet1", "sheet 1") and _is_uf_gen_report_workbook(wb):
            for s in sheetnames:
                if str(s).strip().lower() == "total":
                    return s

        # fallback
        return sheetnames[0] if sheetnames else desired_sheet
    except Exception:
        return desired_sheet


def extract_scada_values_from_exports(
    mapping_rows,
    uploaded_exports: dict,
    file_key_map: dict | None = None,
    target_date=None,
    allow_single_file_fallback: bool = True,
    custom_max_scan_rows: int = 0,
):
    """
    mapping_rows: list[dict] จาก load_scada_excel_mapping
    uploaded_exports: dict filename->bytes ของไฟล์ Excel ที่อัปโหลด
    file_key_map: (optional) dict ของ key_norm -> filename เพื่อบังคับจับคู่ไฟล์ (กันกรณีลูกค้าเปลี่ยนชื่อไฟล์)
    target_date: (optional) datetime.date ที่ผู้ใช้เลือกในหน้า SCADA Export
                 - ถ้าไฟล์มีคอลัมน์ Date (เช่น AF_Report_Gen...) จะใช้กรองให้ตรงวันก่อนเลือกเวลา

    คืนค่า:
      - results: list[dict] สำหรับแสดงในตาราง
      - missing: list[dict] รายการที่ดึงไม่สำเร็จ
    """
    file_key_map = file_key_map or {}

    # ---- lazy workbook cache (กันโหลดไฟล์ใหญ่โดยไม่จำเป็น) ----
    wb_cache: dict[str, openpyxl.Workbook | None] = {}
    wb_is_ufgen: dict[str, bool] = {}

    def get_wb(fname: str):
        if fname in wb_cache:
            return wb_cache[fname]

        b = uploaded_exports.get(fname)
        if b is None:
            wb_cache[fname] = None
            wb_is_ufgen[fname] = False
            return None

        # ไฟล์ใหญ่มาก (เช่น AF_Report) ให้ใช้ read_only เพื่อลด RAM
        read_only = len(b) >= 20_000_000
        try:
            wb = openpyxl.load_workbook(io.BytesIO(b), data_only=True, read_only=read_only)
            wb_cache[fname] = wb
            try:
                wb_is_ufgen[fname] = _is_uf_gen_report_workbook(wb)
            except Exception:
                wb_is_ufgen[fname] = False
            return wb
        except Exception:
            wb_cache[fname] = None
            wb_is_ufgen[fname] = False
            return None

    # helper: หาไฟล์ที่ตรงกับ file_key
    def pick_file_for_key(file_key: str):
        if not uploaded_exports:
            return None

        # normalize key (ตัดวันที่ด้านหน้าออกก่อน เพื่อตรงกับชื่อไฟล์ที่อัปโหลดคนละวัน)
        key_norm = _strip_date_prefix(file_key)
        key_norm2 = _norm_filekey(key_norm)
        key_norm_full = _norm_filekey(file_key)

        fnames = list(uploaded_exports.keys())

        def _strip(fname: str) -> str:
            return _strip_date_prefix(fname)

        def _norm(fname: str) -> str:
            # normalize จากชื่อที่ตัดวันที่แล้ว
            return _norm_filekey(_strip(fname))

        # 0) ถ้าผู้ใช้บังคับ map ไว้ ใช้อันนั้นก่อน
        forced = (
            file_key_map.get(key_norm)
            or file_key_map.get(key_norm2)
            or file_key_map.get(key_norm_full)
        )
        if forced and forced in uploaded_exports:
            return forced

        # 1) match แบบ "ตรงชื่อเป๊ะ" ก่อน (แก้เคส Daily_Report ชนกับ SMMT_Daily_Report)
        if key_norm:
            exact = [f for f in fnames if _strip(f) == key_norm]
            if exact:
                # ถ้า key ไม่ใช่ SMMT ให้เลี่ยงไฟล์ที่มี smmt
                if "smmt" not in key_norm2:
                    non_smmt = [f for f in exact if "smmt" not in _norm(f)]
                    if non_smmt:
                        return non_smmt[0]
                return exact[0]

        if key_norm2:
            exact2 = [f for f in fnames if _norm(f) == key_norm2]
            if exact2:
                if "smmt" not in key_norm2:
                    non_smmt = [f for f in exact2 if "smmt" not in _norm(f)]
                    if non_smmt:
                        return non_smmt[0]
                return exact2[0]

        # 2) UF_System → (สำคัญ) อย่าเปิดไฟล์ทุกตัวเพื่อเดา เพราะไฟล์ใหญ่มากจะช้า
        if "uf_system" in key_norm2 or "ufsystem" in key_norm2:
            for fname in fnames:
                fn = _norm_filekey(fname)
                if "uf_system" in fn or "ufsystem" in fn:
                    return fname
            # fallback: ถ้ามี AF_Report/Report_Gen ให้ใช้แทน UF_System
            for fname in fnames:
                fn = _norm_filekey(fname)
                if "af_report" in fn or "report_gen" in fn or "reportgen" in fn:
                    return fname

        # 3) match แบบ contains + scoring (กรณีชื่อไม่ตรงเป๊ะ)
        def _score(fname: str) -> int:
            s = _strip(fname)
            n = _norm(fname)
            sc = 0
            if key_norm and key_norm in s:
                sc += 6
                if s == key_norm:
                    sc += 10
                if s.endswith(key_norm):
                    sc += 3
            if key_norm2 and key_norm2 in n:
                sc += 6
                if n == key_norm2:
                    sc += 10
                if n.endswith(key_norm2):
                    sc += 3

            # ลงโทษเคสชน SMMT
            if ("smmt" in n) != ("smmt" in key_norm2):
                sc -= 6

            # prefer ใกล้เคียงความยาว (กัน matching กว้างเกิน)
            sc -= abs(len(n) - len(key_norm2))
            return sc

        cand = []
        for fname in fnames:
            s = _strip(fname)
            n = _norm(fname)
            if (key_norm and key_norm in s) or (key_norm2 and key_norm2 in n) or (key_norm_full and key_norm_full in _norm_filekey(fname)):
                cand.append(fname)

        if cand:
            cand.sort(key=_score, reverse=True)
            return cand[0]

        # 4) fallback: ถ้ามีไฟล์เดียว ให้คืนไฟล์นั้น (ปิดได้เพื่อกัน match ผิดตอนประมวลผลไฟล์ใหม่แค่ไฟล์เดียว)
        if allow_single_file_fallback and len(fnames) == 1:
            return fnames[0]

        return None

    # ===== Scan time rows ต่อ sheet แค่ครั้งเดียว =====
    # key ต้องรวม target_date เพราะไฟล์ AF_Report มีหลายวัน
    sheet_ctx_cache = {}  # (fname, sheet, target_date) -> ctx

    import datetime as dt
    from openpyxl.utils.datetime import from_excel

    def _coerce_date(v):
        """แปลงค่า 'วันที่' จากไฟล์ Excel ให้เป็น date

        รองรับหลายแบบเพื่อกันเคสไฟล์ SCADA ใส่วันที่เป็น:
        - datetime / date
        - Excel serial number (เช่น 45291)
        - string (เช่น 2026/01/19, 2026-01-19, 19/01/2026)
        """
        if v is None:
            return None
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v

        # Excel serial date
        if isinstance(v, (int, float)):
            try:
                # บางไฟล์เป็น float เล็ก ๆ ที่ไม่ใช่ serial จริง
                if float(v) > 1:
                    return from_excel(v).date()
            except Exception:
                pass

        # String date
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return None
            # เอาแค่ 10 ตัวแรก เผื่อมีเวลาแนบท้าย
            s10 = s[:10]
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return dt.datetime.strptime(s10, fmt).date()
                except Exception:
                    continue

        return None

    def get_sheet_ctx(fname: str, wb, sheet: str, target_date_local, custom_max_scan_rows: int = 0):
        key = (fname, sheet, target_date_local)
        if key in sheet_ctx_cache:
            return sheet_ctx_cache[key]

        if not wb or sheet not in (wb.sheetnames or []):
            ctx = {"status": "NO_SHEET"}
            sheet_ctx_cache[key] = ctx
            return ctx

        ws = wb[sheet]
        hdr = _find_cell_exact(ws, "Time")
        if not hdr:
            ctx = {"status": "NO_TIME_HEADER"}
            sheet_ctx_cache[key] = ctx
            return ctx

        hdr_row, time_col = hdr

        # หา Date header ที่อยู่แถวเดียวกับ Time (ถ้ามี)
        date_col = None
        try:
            if time_col > 1:
                left = ws.cell(hdr_row, time_col - 1).value
                if isinstance(left, str) and left.strip().lower() == "date":
                    date_col = time_col - 1
            if not date_col:
                # ลองหาในหัวแถวเดียวกัน
                max_c = min(ws.max_column or 0, 40)
                for c in range(1, max_c + 1):
                    v = ws.cell(hdr_row, c).value
                    if isinstance(v, str) and v.strip().lower() == "date":
                        date_col = c
                        break
        except Exception:
            date_col = None

        time_rows: list[tuple[int, int]] = []  # (row_idx, minutes)
        blank_streak = 0

        # ถ้ามี Date column และผู้ใช้เลือกวัน → สแกนจนเจอวันนั้น และหยุดเมื่อเลยวัน (ลดเวลา)
        if date_col and target_date_local:
            started = False
            # กันเคสไฟล์ใหญ่มาก (AF_Report_Gen) ที่ ws.max_row หลอกจนค้าง
            if custom_max_scan_rows > 0:
                max_scan_rows = custom_max_scan_rows
            else:
                max_scan_rows = 50000  # ค่าเริ่มต้น
            max_r = min(ws.max_row or 0, hdr_row + max_scan_rows)
            min_c = min(date_col, time_col)
            max_c = max(date_col, time_col)

            for r, rowvals in enumerate(
                ws.iter_rows(
                    min_row=hdr_row + 1,
                    max_row=max_r,
                    min_col=min_c,
                    max_col=max_c,
                    values_only=True,
                ),
                start=hdr_row + 1,
            ):
                # map ค่าออกมาตามคอลัมน์จริง
                # rowvals จัดตาม min_c..max_c
                def _val_at_col(col):
                    return rowvals[col - min_c]

                dval = _coerce_date(_val_at_col(date_col))
                if dval is None:
                    continue

                if dval < target_date_local:
                    continue

                if dval > target_date_local:
                    if started and time_rows:
                        break
                    continue

                started = True
                tval = _val_at_col(time_col)
                hhmm = _normalize_scada_time(tval)
                mm = _hhmm_to_minutes(hhmm) if hhmm else None
                if mm is not None:
                    time_rows.append((r, mm))
                    blank_streak = 0
                else:
                    blank_streak += 1
                    if blank_streak >= 200 and time_rows:
                        break
        else:
            # ไฟล์ทั่วไป (Daily/SMMT): จำกัด scan ตามค่า custom หรือ 100000 แถว (ไม่จำกัด)
            if custom_max_scan_rows > 0:
                max_scan_rows = custom_max_scan_rows
            else:
                max_scan_rows = 100000  # ค่าเริ่มต้นสแกนเกือบทั้งไฟล์
            max_r = min(ws.max_row or 0, hdr_row + max_scan_rows)

            for r, (tval,) in enumerate(
                ws.iter_rows(
                    min_row=hdr_row + 1,
                    max_row=max_r,
                    min_col=time_col,
                    max_col=time_col,
                    values_only=True,
                ),
                start=hdr_row + 1,
            ):
                hhmm = _normalize_scada_time(tval)
                mm = _hhmm_to_minutes(hhmm) if hhmm else None
                if mm is not None:
                    time_rows.append((r, mm))
                    blank_streak = 0
                else:
                    blank_streak += 1
                    if blank_streak >= 80 and time_rows:
                        break

        if not time_rows:
            ctx = {"status": "NO_DATA_ROW"}
            sheet_ctx_cache[key] = ctx
            return ctx

        ctx = {
            "status": "OK",
            "ws": ws,
            "hdr_row": hdr_row,
            "time_col": time_col,
            "date_col": date_col,
            "time_rows": time_rows,
            "target_row_cache": {},  # hhmm -> row
        }
        sheet_ctx_cache[key] = ctx
        return ctx

    def pick_target_row(ctx, target_time_hhmm: str | None):
        # ถ้าไม่กำหนดเวลา → ใช้แถวสุดท้ายของช่วงที่สแกนได้
        if not target_time_hhmm:
            return ctx["time_rows"][-1][0]

        if target_time_hhmm in ctx["target_row_cache"]:
            return ctx["target_row_cache"][target_time_hhmm]

        tmm = _hhmm_to_minutes(target_time_hhmm)
        row = _find_nearest_time_row(ctx["time_rows"], tmm, max_diff_minutes=300)
        if row is None:
            row = ctx["time_rows"][-1][0]

        ctx["target_row_cache"][target_time_hhmm] = row
        return row

    # ---- สำคัญ: ห้าม ws.cell() กับ read_only workbook เพราะช้ามาก (O(n) ทุกครั้ง) ----
    # จะอ่าน "ทั้งแถว" ด้วย iter_rows แค่ 1 ครั้ง แล้วหยิบค่าคอลัมน์ที่ต้องการ
    row_cache: dict[tuple[str, str, int], tuple] = {}

    results: list[dict] = []
    missing: list[dict] = []

    for row in mapping_rows:
        point_id = row["point_id"]
        file_key = row["file_key"]
        desired_sheet = row.get("sheet") or "Sheet1"
        col = row.get("col") or ""
        t_hhmm = _normalize_scada_time(row.get("time"))

        fname = pick_file_for_key(file_key)
        if not fname:
            missing.append({**row, "reason": "NO_MATCH_FILE"})
            results.append({
                "point_id": point_id,
                "value": None,
                "file": file_key,
                "matched_file": None,
                "sheet": desired_sheet,
                "time": t_hhmm,
                "col": col,
                "status": "NO_FILE",
            })
            continue

        wb = get_wb(fname)
        if not wb:
            missing.append({**row, "reason": "OPEN_FAIL"})
            results.append({
                "point_id": point_id,
                "value": None,
                "file": file_key,
                "matched_file": fname,
                "sheet": desired_sheet,
                "time": t_hhmm,
                "col": col,
                "status": "OPEN_FAIL",
            })
            continue

        sheet = _resolve_sheet_name_for_export(wb, desired_sheet, point_id)
        ctx = get_sheet_ctx(fname, wb, sheet, target_date, custom_max_scan_rows=custom_max_scan_rows)

        if ctx.get("status") != "OK":
            stt = ctx.get("status")
            missing.append({**row, "reason": stt})
            results.append({
                "point_id": point_id,
                "value": None,
                "file": file_key,
                "matched_file": fname,
                "sheet": sheet,
                "time": t_hhmm,
                "col": col,
                "status": stt,
            })
            continue

        # เลือกแถวที่ใกล้เวลาเป้าหมายที่สุด
        target_row = pick_target_row(ctx, t_hhmm)

        # แปลงคอลัมน์ตัวอักษร -> index
        try:
            col_idx = column_index_from_string(str(col).strip().upper())
        except Exception:
            missing.append({**row, "reason": "BAD_COLUMN"})
            results.append({
                "point_id": point_id,
                "value": None,
                "file": file_key,
                "matched_file": fname,
                "sheet": sheet,
                "time": t_hhmm,
                "col": col,
                "status": "BAD_COLUMN",
            })
            continue

        # ดึงทั้งแถวครั้งเดียว (เร็วกว่า ws.cell มาก)
        row_key = (fname, sheet, target_row)
        rowvals = row_cache.get(row_key)
        if rowvals is None:
            try:
                rowvals = next(ctx["ws"].iter_rows(min_row=target_row, max_row=target_row, values_only=True))
                row_cache[row_key] = rowvals
            except StopIteration:
                rowvals = None
            except Exception:
                rowvals = None

        if not rowvals or col_idx > len(rowvals):
            missing.append({**row, "reason": "OUT_OF_RANGE"})
            results.append({
                "point_id": point_id,
                "value": None,
                "file": file_key,
                "matched_file": fname,
                "sheet": sheet,
                "time": t_hhmm,
                "col": col,
                "status": "OUT_OF_RANGE",
            })
            continue

        value = rowvals[col_idx - 1]

        # ทำให้เป็นเลข (ถ้าเป็น string) - ใช้ helper function
        value = parse_scada_numeric_value(value)

        stt = "OK" if value is not None else "EMPTY"
        if stt != "OK":
            missing.append({**row, "reason": stt})

        results.append({
            "point_id": point_id,
            "value": value,
            "file": file_key,
                "matched_file": fname,
            "sheet": sheet,
            "time": t_hhmm,
            "col": col,
            "status": stt,
        })

    return results, missing


def parse_scada_numeric_value(value):
    """
    Parse numeric value from SCADA export ที่อาจมีรูปแบบแตกต่างกัน
    รองรับ: English format (123.45), Thai format (123,45), European format (1.234,56), etc.
    
    Returns:
        float: parsed value, or None if cannot parse
    """
    if value is None:
        return None
    
    # ถ้าเป็น number อยู่แล้ว
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    
    # ถ้าเป็น string
    if isinstance(value, str):
        vv = value.strip()
        
        # Handle empty/invalid strings
        if not vv or vv.lower() in ("", "none", "null", "-", "n/a", "na"):
            return None
        
        # นับจำนวน dots และ commas
        dot_count = vv.count(".")
        comma_count = vv.count(",")
        
        try:
            # Case 1: ไม่มี separator (เช่น "123" หรือ "12345")
            if dot_count == 0 and comma_count == 0:
                return float(vv)
            
            # Case 2: มี dot เดียว (English format เช่น "123.45")
            elif dot_count == 1 and comma_count == 0:
                return float(vv)
            
            # Case 3: มี comma เดียว - ต้องตรวจสอบว่าเป็น decimal หรือ thousands separator
            elif dot_count == 0 and comma_count == 1:
                # ถ้า comma หลังตัวที่ 3 จากท้าย -> น่าจะเป็น thousands separator
                parts = vv.split(",")
                if len(parts[-1]) > 3:
                    # ตัวหลังสุดมากกว่า 3 หลัก -> เป็น decimal แน่ๆ
                    return float(vv.replace(",", "."))
                else:
                    # ตัวหลังสุด <= 3 หลัก -> น่าจะเป็น thousands (เช่น 1,234) แต่อาจเป็น decimal (เช่น 1,5)
                    # ให้ลอง parse แบบ decimal ก่อน ถ้าได้ค่า < 1 ให้ใช้ decimal มิฉะนั้น... ลองนึกใหม่
                    # สำหรับ SCADA โดยทั่วไป: ถ้ามี comma เดียวแล้ว น่าจะเป็น decimal more often
                    return float(vv.replace(",", "."))
            
            # Case 4: มี dot และ comma (thousand separator + decimal)
            elif dot_count == 1 and comma_count == 1:
                last_dot = vv.rfind(".")
                last_comma = vv.rfind(",")
                
                if last_dot > last_comma:
                    # English format with comma thousands: 1,234.56
                    return float(vv.replace(",", ""))
                else:
                    # European format: 1.234,56
                    return float(vv.replace(".", "").replace(",", "."))
            
            # Case 5: หลาย dots, ไม่มี comma (European thousands เช่น "1.234.567")
            elif dot_count > 1 and comma_count == 0:
                return float(vv.replace(".", ""))
            
            # Case 6: หลาย commas, ไม่มี dot
            elif comma_count > 1 and dot_count == 0:
                parts = vv.split(",")
                # เชค: ถ้าตัวหลังสุดมี <= 3 หลักและอย่างน้อย 1 หลัก อาจเป็น decimal
                last_part = parts[-1]
                if 1 <= len(last_part) <= 3:
                    # น่าจะเป็น decimal format (เช่น 1,234,567 with European decimal คือ 1234567.0)
                    # แต่แบบนี้หายากมากสำหรับ SCADA ปกติ
                    # ส่วนใหญ่ commas หลายตัว แปลว่า thousands separator
                    return float(vv.replace(",", ""))
                else:
                    return float(vv.replace(",", ""))
            
            # Case 7: complex (หลาย dots และ commas)
            else:
                # ลองแบบ: remove dots แล้ว replace comma เป็น dot
                temp = vv.replace(".", "").replace(",", ".")
                return float(temp)
        
        except (ValueError, AttributeError):
            return None
    
    return None


def normalize_number_str(s: str, decimals: int = 0) -> str:
    if not s: return ""
    s = str(s).strip().replace(",", "").replace(" ", "")
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"\.{2,}", ".", s)
    if s.count(".") > 1:
        parts = [p for p in s.split(".") if p != ""]
        if len(parts) >= 2: s = parts[0] + "." + "".join(parts[1:])
        else: s = s.replace(".", "")
    if decimals == 0: s = s.replace(".", "")
    return s

# ✅ Template matching สำหรับเลขดิจิทัล (เพิ่ม confidence)
def _create_digit_templates():
    """
    สร้าง template ของ 7-segment display (เลขดิจิทัล 0-9)
    ใช้สำหรับ validate ว่าเลขที่ OCR อ่านมาเป็นเลขดิจิทัลจริง
    """
    # ลายแบบ 7-segment ของเลขแต่ละตัว (แสดงเป็นสัญลักษณ์)
    # segments: a=top, b=top-right, c=bottom-right, d=bottom, e=bottom-left, f=top-left, g=middle
    templates = {
        '0': {'a', 'b', 'c', 'd', 'e', 'f'},        # ทุกด้านยกเว้น middle
        '1': {'b', 'c'},                             # ขวาด้านบนและล่าง
        '2': {'a', 'b', 'd', 'e', 'g'},             # ด้านบน-ขวา-middle-ซ้าย-ล่าง
        '3': {'a', 'b', 'c', 'd', 'g'},             # ด้านบน-ขวา-ล่าง-middle
        '4': {'b', 'c', 'f', 'g'},                  # ขวา-ล่าง-ซ้าย-middle
        '5': {'a', 'c', 'd', 'f', 'g'},             # ด้านบน-ขวา-ล่าง-ซ้าย-middle
        '6': {'a', 'c', 'd', 'e', 'f', 'g'},        # ทุกด้านยกเว้น top-right
        '7': {'a', 'b', 'c'},                        # ด้านบน-ขวา-ล่าง
        '8': {'a', 'b', 'c', 'd', 'e', 'f', 'g'},  # ทั้งหมด
        '9': {'a', 'b', 'c', 'd', 'f', 'g'},        # ทั้งหมดยกเว้น bottom-left
    }
    return templates

_DIGIT_TEMPLATES = _create_digit_templates()

def _validate_digit_char(char_image, digit_template_key: str) -> float:
    """
    ✅ Template matching: ตรวจว่า char_image ตรงกับ template ของเลขตัวใดตัวหนึ่ง
    
    คืนค่า: confidence score 0.0-1.0
    - ตรงเพอร์เฟ็กต์ = 1.0
    - ไม่ตรงสักนิด = 0.7-0.99
    - ไม่ตรงเลย = 0.0-0.5
    """
    if digit_template_key not in _DIGIT_TEMPLATES:
        return 0.5  # unknown digit
    
    try:
        # Convert image เป็น grayscale + binary
        if len(char_image.shape) == 3:
            gray = cv2.cvtColor(char_image, cv2.COLOR_BGR2GRAY)
        else:
            gray = char_image
        
        _, binary = cv2.threshold(gray, 128, 255, cv2.THRESH_BINARY)
        
        # ตัวอักษร 7-segment แบ่งเป็น 7 ส่วน (ประมาณ)
        h, w = binary.shape
        
        # ตัดแต่ละ segment (approximation)
        segment_regions = {
            'a': (0, 0, w, h//4),           # top
            'b': (w//2, h//4, w, h//2),    # top-right
            'c': (w//2, h//2, w, 3*h//4),  # bottom-right
            'd': (0, 3*h//4, w, h),        # bottom
            'e': (0, h//2, w//2, 3*h//4),  # bottom-left
            'f': (0, h//4, w//2, h//2),    # top-left
            'g': (w//4, h//2-5, 3*w//4, h//2+5),  # middle
        }
        
        detected_segments = set()
        for seg_name, (x1, y1, x2, y2) in segment_regions.items():
            region = binary[max(0, y1):min(h, y2), max(0, x1):min(w, x2)]
            if region.size > 0:
                ratio = np.sum(region > 128) / region.size
                if ratio > 0.15:  # ถ้า >15% สว่าง ถือว่ามี segment นี้
                    detected_segments.add(seg_name)
        
        template = _DIGIT_TEMPLATES[digit_template_key]
        
        # คำนวณ Jaccard similarity
        intersection = len(detected_segments & template)
        union = len(detected_segments | template)
        
        if union == 0:
            return 0.5
        
        confidence = intersection / union
        return max(0.0, min(1.0, confidence))
    
    except Exception:
        return 0.5

def _apply_template_matching_refinement(candidates: list, decimals: int = 0) -> list:
    """
    ✅ Refine candidates โดยใช้ Template Matching
    
    - เอาแต่เลขที่ดูเหมือน 7-segment display
    - ให้ confidence score เพิ่ม/ลด ตามว่า match template ไหม
    """
    refined = []
    
    for c in candidates:
        try:
            val_str = str(c.get("val", ""))
            score = float(c.get("score", 0))
            
            # Sum template confidence สำหรับแต่ละ digit
            template_score = 0.0
            for digit in val_str.replace(".", ""):
                if digit.isdigit():
                    conf = _validate_digit_char(np.zeros((20, 20), dtype=np.uint8), digit)
                    template_score += conf
            
            # ค่าเฉลี่ย
            if len(val_str.replace(".", "")) > 0:
                template_score = template_score / len(val_str.replace(".", ""))
            else:
                template_score = 0.5
            
            # Combine: original score 70% + template score 30%
            refined_score = (score * 0.7) + (template_score * 100 * 0.3)  # scale template_score
            
            refined_c = dict(c)
            refined_c["combined_template_score"] = refined_score
            refined.append(refined_c)
        except Exception:
            refined.append(c)
    
    # Sort by refined score
    refined.sort(key=lambda x: x.get("combined_template_score", 0), reverse=True)
    return refined

def preprocess_text(text):
    patterns = [r'IP\s*51', r'50\s*Hz', r'Class\s*2', r'3x220/380\s*V', r'Type', r'Mitsubishi', r'Electric', r'Wire', r'kWh', r'MH\s*[-]?\s*96', r'30\s*\(100\)\s*A', r'\d+\s*rev/kWh', r'WATT-HOUR\s*METER', r'Indoor\s*Use', r'Made\s*in\s*Thailand']
    for p in patterns: text = re.sub(p, '', text, flags=re.IGNORECASE)
    text = re.sub(r'\b10,000\b', '', text)
    text = re.sub(r'\b1,000\b', '', text)
    text = re.sub(r'(?<=[\d\s])[\|Il!](?=[\d\s])', '1', text)
    text = re.sub(r'(?<=[\d\s])[Oo](?=[\d\s])', '0', text)
    return text

def is_digital_meter(config):
    blob = f"{config.get('type','')} {config.get('name','')} {config.get('keyword','')}".lower()
    return ("digital" in blob) or ("scada" in blob) or (int(config.get('decimals', 0) or 0) > 0)

def preprocess_image_cv(image_bytes, config, use_roi=True, variant="auto"):
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
    if img is None: return image_bytes

    H, W = img.shape[:2]
    if W > 1280:
        scale = 1280 / W
        img = cv2.resize(img, (1280, int(H * scale)), interpolation=cv2.INTER_AREA)
        H, W = img.shape[:2]

    if use_roi:
        x1, y1, x2, y2 = config.get('roi_x1', 0), config.get('roi_y1', 0), config.get('roi_x2', 0), config.get('roi_y2', 0)
        if x2 and y2:
            if 0 < x2 <= 1 and 0 < y2 <= 1:
                x1, y1, x2, y2 = int(float(x1) * W), int(float(y1) * H), int(float(x2) * W), int(float(y2) * H)
            else:
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            pad_x, pad_y = int(0.03 * W), int(0.03 * H)
            x1, y1 = max(0, x1 - pad_x), max(0, y1 - pad_y)
            x2, y2 = min(W, x2 + pad_x), min(H, y2 + pad_y)
            if x2 > x1 and y2 > y1:
                img = img[y1:y2, x1:x2]
                H, W = img.shape[:2]

    if config.get('ignore_red', False):
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        # ✅ เข้มงวดมาก: ตัดเฉพาะเลขแดงเจาะจง
        # Red range ที่เข้มงวด: H=[0,10] หรือ [170,180] + S>=85 + V>=70
        lower_red1 = np.array([0, 85, 70])
        upper_red1 = np.array([10, 255, 255])
        lower_red2 = np.array([170, 85, 70])
        upper_red2 = np.array([180, 255, 255])
        
        mask_red = cv2.inRange(hsv, lower_red1, upper_red1) + cv2.inRange(hsv, lower_red2, upper_red2)
        
        # ✅ Morphological operations: ทำให้ mask ติดกัน
        kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        mask_red = cv2.morphologyEx(mask_red, cv2.MORPH_CLOSE, kernel, iterations=2)
        mask_red = cv2.morphologyEx(mask_red, cv2.MORPH_OPEN, kernel, iterations=1)
        
        # ✅ เปลี่ยนเลขแดงเป็นสีขาว (255,255,255)
        img[mask_red > 0] = [255, 255, 255]

    if variant == "raw":
        ok, encoded = cv2.imencode(".jpg", img)
        return encoded.tobytes() if ok else image_bytes

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    if variant == "invert": gray = 255 - gray

    use_digital_logic = (variant == "soft") or (variant == "auto" and is_digital_meter(config))
    # ✅ Analog meter with ignore_red should also use enhanced preprocessing
    use_enhanced_analog = (variant == "auto" and not is_digital_meter(config) and config.get('ignore_red', False))

    if use_digital_logic or use_enhanced_analog:
        if min(H, W) < 300:
            gray = cv2.resize(gray, None, fx=2.5, fy=2.5, interpolation=cv2.INTER_CUBIC)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        g = clahe.apply(gray)
        blur = cv2.GaussianBlur(g, (0, 0), 1.0)
        sharp = cv2.addWeighted(g, 1.6, blur, -0.6, 0)
        
        # ✅ เพิ่ม Morphological operations เพื่อชาร์ปเลขดิจิทัล
        # 1) Bilateral filter: ลดสัญญาณรบกวนแต่เก็บขอบไว้
        sharp = cv2.bilateralFilter(sharp, 5, 50, 50)
        
        # 2) Morphological closing: เติมหลุมเล็ก ๆ ในตัวเลข (ทำให้ตัวเลขแน่นขึ้น)
        kernel_close = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
        sharp = cv2.morphologyEx(sharp, cv2.MORPH_CLOSE, kernel_close, iterations=1)
        
        # 3) Morphological opening: ลบสัญญาณรบกวนเล็ก ๆ (ตัดเสี่ยวนอก)
        kernel_open = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2, 2))
        sharp = cv2.morphologyEx(sharp, cv2.MORPH_OPEN, kernel_open, iterations=1)
        
        # 4) Dilate: ขยายเลขให้ฟูกว่างขึ้นเล็กน้อยเพื่อให้ OCR เห็นชัด
        kernel_dilate = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (2, 2))
        sharp = cv2.dilate(sharp, kernel_dilate, iterations=1)
        
        # 5) Final sharpening: ชาร์พขอบเลขให้แหลมขึ้น
        kernel_sharpen = np.array([[-1, -1, -1],
                                   [-1,  9, -1],
                                   [-1, -1, -1]])
        sharp = cv2.filter2D(sharp, -1, kernel_sharpen)
        sharp = np.clip(sharp, 0, 255).astype(np.uint8)
        
        ok, encoded = cv2.imencode(".png", sharp)
        return encoded.tobytes() if ok else image_bytes
    else:
        gray2 = cv2.bilateralFilter(gray, 7, 50, 50)
        th = cv2.adaptiveThreshold(gray2, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 7)
        ok, encoded = cv2.imencode(".png", th)
        return encoded.tobytes() if ok else image_bytes

def _roboflow_detect_digits(image_bytes, api_key=None, model_id="water-meter-monitoring/1"):
    """
    🔥 ใช้ Roboflow YOLO Model ตรวจจับตัวเลขบนมาตรวัดน้ำแบบ Object Detection
    - แม่นกว่า OCR ธรรมดาเพราะ train มาเจาะจง
    - ได้ bounding box + confidence score แต่ละหลัก
    - เรียงตามตำแหน่ง x (จากซ้ายไปขวา)
    
    คืนค่า: (digit_sequence: str, confidence: float, predictions: list) หรือ (None, 0, [])
    """
    if not HAS_ROBOFLOW:
        print("⚠️ Roboflow SDK ไม่ได้ติดตั้ง")
        return None, 0, []
    
    if not api_key:
        # ลองอ่านจาก secrets (ถ้ามี)
        try:
            api_key = st.secrets.get('roboflow_api_key', None)
        except:
            api_key = None
    
    if not api_key:
        print("⚠️ ไม่พบ roboflow_api_key → ใช้ Vision OCR แทน")
        return None, 0, []
    
    try:
        # Initialize client
        client = InferenceHTTPClient(
            api_url="https://detect.roboflow.com",
            api_key=api_key
        )
        
        # เซฟรูปชั่วคราวเพื่อส่งให้ Roboflow (ต้องการ file path หรือ bytes)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            tmp.write(image_bytes)
            tmp_path = tmp.name
        
        try:
            # Run inference
            print("🔥 กำลังใช้ Roboflow AI ตรวจจับตัวเลข...")
            result = client.infer(tmp_path, model_id=model_id)
            predictions = result.get('predictions', [])
            
            if not predictions:
                print("⚠️ Roboflow ไม่พบตัวเลข → fallback to OCR")
                return None, 0, []
            
            # เรียงตาม x coordinate (จากซ้ายไปขวา)
            sorted_preds = sorted(predictions, key=lambda p: p.get('x', 0))
            
            # สร้างเลขจาก class labels
            digit_sequence = ''.join([str(pred.get('class', '')) for pred in sorted_preds])
            
            # คำนวณ average confidence
            avg_confidence = sum([pred.get('confidence', 0) for pred in sorted_preds]) / len(sorted_preds)
            
            print(f"✅ Roboflow: ตรวจพบ '{digit_sequence}' (confidence: {avg_confidence:.2%}, หลัก: {len(sorted_preds)})")
            return digit_sequence, avg_confidence, sorted_preds
        
        finally:
            # ลบไฟล์ชั่วคราว
            try:
                os.remove(tmp_path)
            except:
                pass
    
    except Exception as e:
        # Log error แล้ว fallback to Vision OCR
        print(f"⚠️ Roboflow error: {str(e)[:100]} → fallback to OCR")
        return None, 0, []

def _vision_read_text(processed_bytes):
    try:
        image = vision.Image(content=processed_bytes)
        ctx = vision.ImageContext(language_hints=["en"])
        resp = VISION_CLIENT.text_detection(image=image, image_context=ctx)
        if getattr(resp, "error", None) and resp.error.message: return "", resp.error.message
        if resp.text_annotations: return (resp.text_annotations[0].description or ""), ""
        
        resp2 = VISION_CLIENT.document_text_detection(image=image, image_context=ctx)
        txt = ""
        if resp2.full_text_annotation and resp2.full_text_annotation.text: txt = resp2.full_text_annotation.text
        return (txt or ""), ""
    except Exception as e:
        return "", str(e)

def ocr_process(image_bytes, config, debug=False, return_candidates=False, use_roboflow=True):
    decimal_places = int(config.get('decimals', 0) or 0)
    keyword = str(config.get('keyword', '') or '').strip()
    expected_digits = int(config.get('expected_digits', 0) or 0)
    
    # 🔥 ลอง Roboflow Detection ก่อน (แม่นกว่า OCR สำหรับ water meter)
    if use_roboflow and HAS_ROBOFLOW:
        try:
            # ใช้รูปที่ preprocessed ROI (ถ้ามี)
            roi_bytes = image_bytes
            if config.get('roi_x1') and config.get('roi_x2'):
                roi_bytes = preprocess_image_cv(image_bytes, config, use_roi=True, variant="raw")
            
            digit_seq, confidence, preds = _roboflow_detect_digits(roi_bytes)
            
            if digit_seq and confidence > 0.6:  # threshold ความมั่นใจ
                # แปลง digit_sequence เป็น float ตาม decimal_places
                try:
                    val_str = str(digit_seq).strip()
                    
                    # ถ้ามีจุดทศนิยมใน sequence แล้ว
                    if '.' in val_str:
                        val = float(val_str)
                    else:
                        # ใส่จุดทศนิยมตาม config
                        if decimal_places > 0 and len(val_str) > decimal_places:
                            int_part = val_str[:-decimal_places]
                            dec_part = val_str[-decimal_places:]
                            val = float(f"{int_part}.{dec_part}")
                        else:
                            val = float(val_str)
                    
                    # Validate ตาม expected_digits
                    if expected_digits > 0:
                        digit_count = len(str(int(abs(val))))
                        if digit_count > expected_digits + 1:  # เกินจริง ๆ → ไม่ใช้
                            raise ValueError("Too many digits")
                    
                    # 🎯 Return ทันทีถ้า Roboflow สำเร็จ
                    print(f"🎯 ใช้ผลลัพธ์ Roboflow: {val} (raw: {val_str}, conf: {confidence:.2%})")
                    if return_candidates:
                        candidates = [{
                            "val": val,
                            "score": confidence * 1000,  # scale to match OCR scoring
                            "raw": val_str,
                            "method": "roboflow",
                            "confidence": confidence,
                            "predictions": len(preds)
                        }]
                        return val, candidates
                    else:
                        return val
                
                except (ValueError, ZeroDivisionError) as e:
                    print(f"⚠️ Roboflow parsing error: {e} → fallback to OCR")
        
        except Exception as e:
            print(f"⚠️ Roboflow exception: {str(e)[:100]} → fallback to OCR")
    
    # ✅ Fallback: ใช้ Vision OCR แบบเดิม
    print("🔄 ใช้ Vision OCR (Google Cloud Vision)...")
    attempts = [
        ("ROI_auto",  True,  "auto"),
        ("ROI_raw",   True,  "raw"),
        ("ROI_soft",  True,  "soft"),
        ("ROI_inv",   True,  "invert"),
        ("FULL_auto", False, "auto"),
        ("FULL_raw",  False, "raw"),
    ]

    def _has_digits(s: str) -> bool:
        return bool(s) and any(c.isdigit() for c in s)

    def check_digits_len(val: float) -> int:
        """คืนค่า 'จำนวนหลัก' ของเลขหน้า (ไม่รวมทศนิยม)"""
        try:
            return len(str(int(abs(float(val)))))
        except Exception:
            return 0

    def check_digits_ok(val: float) -> bool:
        """อนุญาตให้ผ่านแบบ 'ยืดหยุ่น' แต่จะไปจัดการด้วยคะแนนอีกที"""
        if val is None:
            return False
        # ❌ มิเตอร์ไม่ควรติดลบ
        if float(val) < 0:
            return False
        if expected_digits <= 0:
            return True
        ln = check_digits_len(val)
        # ยังยอมให้ +1 ได้ (กันเคสเลขโตขึ้นจริง) แต่จะโดนหักคะแนนหนัก
        return 1 <= ln <= expected_digits + 1

    def looks_like_spec_context(text: str, start: int, end: int) -> bool:
        """ดูรอบ ๆ ตัวเลขว่าเป็นเลขสเปคเครื่อง (Hz/V/A/IP/Rev) ไหม"""
        ctx = text[max(0, start - 12):min(len(text), end + 12)].lower()
        # ถ้าใกล้ ๆ มี kWh ให้ถือว่าไม่ใช่สเปค (เป็นเลขมิเตอร์ได้)
        if "kwh" in ctx or "kw h" in ctx:
            return False
        bad = ["hz", "volt", " v", "v ", "amp", " a", "a ", "class", "ip", "rev", "rpm", "phase", "3x", "indoor"]
        return any(b in ctx for b in bad)

    common_noise = {10, 30, 50, 60, 100, 220, 230, 240, 380, 400, 415, 1000, 10000}

    best_val = None
    best_score = -10**9

    # ✅ เก็บ candidate ข้าม attempts (ไว้ใช้ History Guard)
    all_candidates = []
    TOPK = 60  # กัน list โตเกิน
    
    for tag, use_roi, variant in attempts:
        processed = preprocess_image_cv(image_bytes, config, use_roi=use_roi, variant=variant)
        txt, err = _vision_read_text(processed)
        if not txt or not _has_digits(txt):
            continue
            
        raw_text = (txt or "").replace("\n", " ")
        raw_text = re.sub(r"\.{2,}", ".", raw_text)

        # preprocess แล้วใช้ "ข้อความเดียวกัน" ในการหาเลข + เช็คบริบท (แก้บั๊กตำแหน่ง)
        full_text = preprocess_text(raw_text)
        full_text = re.sub(r"\.{2,}", ".", full_text)

        # ตัดปีออกแบบไม่ทำให้ตำแหน่งพัง (จะใช้ full_text ที่ตัดแล้วทั้งคู่)
        scan_text = re.sub(r"\b202[0-9]\b|\b256[0-9]\b", "", full_text)
        scan_text = re.sub(r"\.{2,}", ".", scan_text)

        candidates = []

        # ---- โบนัสตาม attempt (กัน FULL ภาพกว้างชนะ ROI ง่ายเกิน) ----
        attempt_bonus = 0
        if use_roi:
            attempt_bonus += 80
        if variant in ("soft", "auto"):
            attempt_bonus += 10
        
        # ---- 1) ลองจับจาก keyword ก่อน (แม่นสุด) ----
        if keyword:
            kw = re.escape(keyword)
            patterns = [
                kw + r"[^\d]*((?:\d|O|o|l|I|\|)+[\.,]?\d*)",
                r"((?:\d|O|o|l|I|\|)+[\.,]?\d*)[^\d]*" + kw
            ]
            for pat in patterns:
                match = re.search(pat, raw_text, re.IGNORECASE)
                if match:
                    val_str = match.group(1)
                    val_str = val_str.replace("O", "0").replace("o", "0").replace("l", "1").replace("I", "1").replace("|", "1")
                    val_str = normalize_number_str(val_str, decimal_places)
                    try:
                        val = float(val_str)
                        if decimal_places > 0 and "." not in val_str:
                            val = val / (10 ** decimal_places)
                        if check_digits_ok(val):
                            score = 900 + attempt_bonus  # ให้สูงมาก เพราะเจอ keyword
                            ln = check_digits_len(val)

                            # ให้คะแนน "ใกล้ expected_digits" ชนะ (ไม่ใช่เลขยาวชนะ)
                            if expected_digits > 0:
                                score += max(0, 160 - abs(ln - expected_digits) * 60)
                                if ln == expected_digits:
                                    score += 80
                                if ln == expected_digits + 1:
                                    score -= 80  # หักหนักกรณี +1
                            candidates.append({"val": float(val), "score": score})
                    except Exception:
                            pass

        # ---- 2) กวาดเลขทั้งหมดในข้อความ ----
        for m in re.finditer(r"(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?", scan_text):
            n_str = m.group(0)

            # เช็คว่าเลขนี้เป็นสเปคเครื่องไหม (ใช้ scan_text ตัวเดียวกัน)
            if looks_like_spec_context(scan_text, m.start(), m.end()):
                continue

            n_str2 = normalize_number_str(n_str, decimal_places)
            if not n_str2:
                continue
            
            try:
                val = float(n_str2) if "." in n_str2 else float(int(n_str2))
                if decimal_places > 0 and "." not in n_str2:
                    val = val / (10 ** decimal_places)

                # ❌ ไม่เอาติดลบ
                if float(val) < 0:
                    continue

                # กันเลข noise ยอดฮิต ถ้าไม่มี keyword ช่วย
                if int(abs(val)) in common_noise and not keyword:
                    continue

                if not check_digits_ok(val):
                    continue

                ln = check_digits_len(val)
                score = 200 + attempt_bonus

                # ให้คะแนน "ใกล้ expected_digits" ชนะ
                if expected_digits > 0:
                    score += max(0, 140 - abs(ln - expected_digits) * 50)
                    if ln == expected_digits:
                        score += 60
                    if ln == expected_digits + 1:
                        score -= 70
                else:
                    # ถ้าไม่กำหนด expected_digits ให้พอใช้ logic เดิม (เบา ๆ)
                    score += min(ln, 10) * 6

                # ถ้ามีทศนิยมและต้องการทศนิยม ให้บวกนิดหน่อย
                if decimal_places > 0 and "." in n_str2:
                    score += 20

                candidates.append({"val": float(val), "score": score, "tag": tag})
            except Exception:
                continue
            
        if candidates:
            pick = max(candidates, key=lambda x: x["score"])
            if pick["score"] > best_score:
                best_score = pick["score"]
                best_val = pick["val"]
            # ✅ รวม candidates ข้าม attempts (เก็บเฉพาะ topK)
            all_candidates.extend(candidates)
            all_candidates.sort(key=lambda x: float(x.get("score", 0)), reverse=True)
            if len(all_candidates) > TOPK:
                all_candidates = all_candidates[:TOPK]
                
            # ถ้าเจอคะแนนสูงมากแล้ว ก็พอ (กันเรียก Vision หลายรอบ)
            if best_score >= 980:
                break

    final_val = float(best_val) if best_val is not None else 0.0
    
    # ✅ dedupe candidates ตามค่า val (เก็บคะแนนดีที่สุด)
    if return_candidates:
        by_val = {}
        for c in all_candidates:
            try:
                v = float(c.get("val"))
            except Exception:
                continue
            key = round(v, max(0, decimal_places) + 2)
            if key not in by_val or float(c.get("score", 0)) > float(by_val[key].get("score", 0)):
                by_val[key] = {"val": v, "score": float(c.get("score", 0)), "tag": c.get("tag", "")}
        
        cand_out = sorted(by_val.values(), key=lambda x: x["score"], reverse=True)[:25]
        return final_val, cand_out
        
    return final_val
         
# =========================================================
# --- 🔳 QR + REF IMAGE HELPERS (Mobile) ---
# =========================================================
REF_IMAGE_FOLDER = "ref_images"  # โฟลเดอร์รูปตัวอย่างใน Bucket

def get_ref_image_url(point_id: str) -> str:
    pid = str(point_id).strip().upper()
    return f"https://storage.googleapis.com/{BUCKET_NAME}/{REF_IMAGE_FOLDER}/{pid}.jpg"

def decode_qr(image_bytes: bytes):
    """คืนค่า point_id จาก QR (ถ้าอ่านไม่ได้จะคืน None)"""
    try:
        arr = np.frombuffer(image_bytes, np.uint8)
        img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if img is None:
            return None
        detector = cv2.QRCodeDetector()
        data, _, _ = detector.detectAndDecode(img)
        data = (data or "").strip()
        return data.upper() if data else None
    except:
        return None

def infer_meter_type(config: dict) -> str:
    """เดา meter_type จาก config เพื่อกันกรอกผิด"""
    blob = f"{config.get('type','')} {config.get('name','')}".lower()
    if ("น้ำ" in blob) or ("water" in blob) or ("ประปา" in blob):
        return "Water"
    return "Electric"
# =========================================================
# --- 🖥️ DASHBOARD SCREENSHOT OCR (FLOW 1-3) ---
# =========================================================

# ปรับ point_id default ให้ตรงกับ PointsMaster ของคุณได้เลย
_DASH_DEFAULT_POINT_MAP = {
    # FLOW 1
    (1, "pressure_bar"): "C_Bar_FLOW_1",
    (1, "flowrate_m3h"): "D_m_h_FLOW_1",
    (1, "flow_total_m3"): "J_FLOW_1",
    # FLOW 2
    (2, "pressure_bar"): "E_Bar_FLOW_2",
    (2, "flowrate_m3h"): "F_m_h_FLOW_2",
    (2, "flow_total_m3"): "K_FLOW_2",
    # FLOW 3
    (3, "pressure_bar"): "G_Bar_FLOW_3",
    (3, "flowrate_m3h"): "H_m_h_FLOW_3",
    (3, "flow_total_m3"): "L_FLOW_3",
}

_NUM_RE = re.compile(r"^[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^[-+]?\d+(?:\.\d+)?$")

def _looks_like_number(s: str) -> bool:
    if s is None:
        return False
    s = str(s).strip()
    if not s:
        return False
    if ":" in s or "/" in s:
        return False
    s2 = s.replace("O", "0").replace("o", "0")
    return bool(_NUM_RE.match(s2))

def _parse_number(s: str):
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    s = s.replace("O", "0").replace("o", "0").replace(",", "")
    try:
        return float(s)
    except Exception:
        return None

def _cv2_decode_bytes(image_bytes: bytes):
    arr = np.frombuffer(image_bytes, np.uint8)
    return cv2.imdecode(arr, cv2.IMREAD_COLOR)

def _cv2_encode_jpg(img, quality: int = 92) -> bytes:
    ok, buf = cv2.imencode('.jpg', img, [int(cv2.IMWRITE_JPEG_QUALITY), int(quality)])
    return buf.tobytes() if ok else b""

def _upscale_for_ocr(img, max_side: int = 2200):
    if img is None:
        return img
    h, w = img.shape[:2]
    scale = 2.0
    if max(h, w) * scale > max_side:
        scale = max_side / float(max(h, w))
    if scale <= 1.05:
        return img
    return cv2.resize(img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)

def _vision_tokens(image_bytes: bytes, lang_hints=("en",)):
    image = vision.Image(content=image_bytes)
    ctx = vision.ImageContext(language_hints=list(lang_hints))
    resp = VISION_CLIENT.text_detection(image=image, image_context=ctx)
    if resp.error.message:
        raise RuntimeError(resp.error.message)

    ann = resp.text_annotations
    tokens = []
    for a in ann[1:]:
        txt = (a.description or "").strip()
        if not txt:
            continue
        vs = a.bounding_poly.vertices
        xs = [v.x for v in vs]
        ys = [v.y for v in vs]
        x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
        tokens.append({
            "text": txt,
            "x1": x1, "y1": y1, "x2": x2, "y2": y2,
            "cx": (x1 + x2) / 2.0,
            "cy": (y1 + y2) / 2.0,
            "h": max(1.0, (y2 - y1)),
        })
    full_text = ann[0].description if ann else ""
    return tokens, full_text

def _norm_token_text(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())

def _suggest_dashboard_crop(tokens, w: int, h: int):
    # default crop: ตัด sidebar + top bar
    def_roi = (int(w * 0.18), int(h * 0.18), int(w * 0.99), int(h * 0.92))
    if not tokens:
        return def_roi

    anchors = []
    for t in tokens:
        tn = _norm_token_text(t.get("text", ""))
        if any(k in tn for k in ["FLOW", "PRESSURE", "FLOWRATE", "FLOWTOTAL", "TOTALM3", "M3H", "BAR"]):
            anchors.append(t)

    if not anchors:
        return def_roi

    x1 = min(t["x1"] for t in anchors)
    y1 = min(t["y1"] for t in anchors)
    x2 = max(t["x2"] for t in anchors)
    y2 = max(t["y2"] for t in anchors)

    pad_x_left  = int(0.05 * w)
    pad_x_right = int(0.35 * w)
    pad_y_top   = int(0.10 * h)
    pad_y_bot   = int(0.45 * h)

    rx1 = max(0, x1 - pad_x_left)
    ry1 = max(0, y1 - pad_y_top)
    rx2 = min(w, x2 + pad_x_right)
    ry2 = min(h, y2 + pad_y_bot)

    if (rx2 - rx1) < int(0.35 * w) or (ry2 - ry1) < int(0.20 * h):
        return def_roi
    return (rx1, ry1, rx2, ry2)

def _join_adjacent_numeric_tokens(num_tokens, gap_px: int = 14):
    if not num_tokens:
        return []
    num_tokens = sorted(num_tokens, key=lambda t: t["x1"])
    merged = []
    cur = dict(num_tokens[0])
    for t in num_tokens[1:]:
        gap = t["x1"] - cur["x2"]
        if gap >= 0 and gap <= gap_px:
            cur["text"] = f"{cur['text']}{t['text']}"
            cur["x2"] = max(cur["x2"], t["x2"])
            cur["y1"] = min(cur.get("y1", 0), t.get("y1", 0))
            cur["y2"] = max(cur.get("y2", 0), t.get("y2", 0))
        else:
            merged.append(cur)
            cur = dict(t)
    merged.append(cur)

    for m in merged:
        m["cx"] = (m["x1"] + m["x2"]) / 2.0
        m["cy"] = (m["y1"] + m["y2"]) / 2.0
        m["h"] = max(1.0, (m["y2"] - m["y1"]))
    return merged

def extract_dashboard_flow_values(image_bytes: bytes, debug: bool = False):
    """อ่านค่า FLOW 1-3 จากภาพ Dashboard
    คืนค่า list[dict]: flow, pressure_bar, flowrate_m3h, flow_total_m3, status
    """
    img = _cv2_decode_bytes(image_bytes)
    if img is None:
        rows = [{"flow": f"FLOW {i}", "pressure_bar": None, "flowrate_m3h": None, "flow_total_m3": None, "status": "BAD_IMAGE"} for i in (1,2,3)]
        return (rows, {"reason": "cv2_decode_failed"}) if debug else rows

    h, w = img.shape[:2]

    # pass1: full OCR เพื่อหา ROI
    try:
        tokens1, full_text1 = _vision_tokens(image_bytes, lang_hints=("en",))
    except Exception as e:
        rows = [{"flow": f"FLOW {i}", "pressure_bar": None, "flowrate_m3h": None, "flow_total_m3": None, "status": "VISION_ERROR"} for i in (1,2,3)]
        return (rows, {"error": str(e)}) if debug else rows

    x1, y1, x2, y2 = _suggest_dashboard_crop(tokens1, w, h)
    crop = img[y1:y2, x1:x2].copy()
    crop = _upscale_for_ocr(crop)
    crop_bytes = _cv2_encode_jpg(crop, quality=92)

    # pass2: OCR บน crop
    try:
        tokens, full_text = _vision_tokens(crop_bytes, lang_hints=("en",))
    except Exception:
        tokens, full_text = tokens1, full_text1

    flow_rows = {}
    for t in tokens:
        tn = _norm_token_text(t.get("text", ""))
        m = re.match(r"^FLOW([123])$", tn)
        if m:
            n = int(m.group(1))
            flow_rows[n] = {"y": t["cy"], "h": t["h"], "x_right": t["x2"]}

    # FLOW + digit แยกกัน
    if len(flow_rows) < 3:
        flow_tokens = [t for t in tokens if _norm_token_text(t.get("text","")) == "FLOW"]
        digit_tokens = [t for t in tokens if str(t.get("text","")).strip() in ("1","2","3")]
        for d in digit_tokens:
            n = int(str(d["text"]))
            if n in flow_rows:
                continue
            best = None
            best_score = 1e9
            for f in flow_tokens:
                dx = abs(d["cx"] - f["cx"])
                dy = abs(d["cy"] - f["cy"])
                score = dx + dy * 1.2
                if score < best_score and dx < 120 and dy < 120:
                    best, best_score = f, score
            if best:
                y = (best["cy"] + d["cy"]) / 2.0
                hh = max(best["h"], d["h"]) * 1.8
                xr = max(best["x2"], d["x2"])
                flow_rows[n] = {"y": y, "h": hh, "x_right": xr}

    out_rows = []
    for n in (1,2,3):
        row = {"flow": f"FLOW {n}", "pressure_bar": None, "flowrate_m3h": None, "flow_total_m3": None, "status": "NOT_FOUND"}
        meta = flow_rows.get(n)
        if not meta:
            out_rows.append(row)
            continue

        band = max(22.0, meta["h"] * 1.2)
        x_min = meta["x_right"] + 8

        row_tokens = [t for t in tokens if abs(t["cy"] - meta["y"]) <= band and t["x1"] >= x_min]
        num_tokens = [t for t in row_tokens if _looks_like_number(t.get("text",""))]
        num_tokens = _join_adjacent_numeric_tokens(num_tokens, gap_px=14)
        num_tokens = [t for t in num_tokens if _looks_like_number(t.get("text",""))]
        num_tokens = sorted(num_tokens, key=lambda t: t["cx"])

        if len(num_tokens) >= 3:
            p   = _parse_number(num_tokens[0]["text"])
            fr  = _parse_number(num_tokens[1]["text"])
            # ใช้ num_tokens[3] (TOT1) ถ้ามี 4 ตัวเลขขึ้นไป ไม่เช่นนั้นใช้ num_tokens[2]
            tot = _parse_number(num_tokens[3]["text"]) if len(num_tokens) >= 4 else _parse_number(num_tokens[2]["text"])
            row.update({
                "pressure_bar": p,
                "flowrate_m3h": fr,
                "flow_total_m3": tot,
                "status": "OK" if (p is not None and fr is not None and tot is not None) else "PARTIAL",
            })

        out_rows.append(row)

    dbg = {
        "roi": {"x1": x1, "y1": y1, "x2": x2, "y2": y2},
        "flow_rows": flow_rows,
        "full_text": (full_text or "")[:2000],
        "full_text_pass1": (full_text1 or "")[:2000],
        "tokens_count": len(tokens),
    }
    return (out_rows, dbg) if debug else out_rows

# =========================================================
# --- 📸 BULK IMAGE OCR: FIND point_id FROM PHOTO ---
# =========================================================

def _norm_pid_key(s: str) -> str:
    s = str(s or "").upper().strip()
    s = s.replace("-", "_")
    s = re.sub(r"\s+", "_", s)          # space -> _
    s = re.sub(r"[^A-Z0-9_]", "", s)    # ตัดสัญลักษณ์แปลกๆ
    s = re.sub(r"_+", "_", s).strip("_")
    return s

@st.cache_data(ttl=3600)
def build_pid_norm_map():
    """สร้าง map สำหรับ match point_id แบบทน OCR เพี้ยน"""
    pm = load_points_master() or []
    norm_map = {}
    for r in pm:
        pid = str(r.get("point_id", "")).strip().upper()
        if not pid:
            continue
        norm_map[_norm_pid_key(pid)] = pid
    return norm_map

def _crop_bottom_bytes(image_bytes: bytes, frac: float = 0.40) -> bytes:
    """ครอปช่วงล่างของรูป (ตรงเทปเหลือง) เพื่อ OCR point_id ให้แม่น/เร็ว"""
    img = _cv2_decode_bytes(image_bytes)
    if img is None:
        return image_bytes
    h, w = img.shape[:2]
    y1 = int(h * (1.0 - frac))
    crop = img[y1:h, 0:w].copy()
    crop = _upscale_for_ocr(crop, max_side=2200)
    out = _cv2_encode_jpg(crop, quality=92)
    return out or image_bytes

def find_point_id_from_text(ocr_text: str, norm_map: dict):
    t = _norm_pid_key(ocr_text)
    if not t:
        return None

    # 1) exact substring match (เร็ว+แม่น)
    best = None
    best_len = -1
    for nkey, orig in norm_map.items():
        if nkey and nkey in t:
            if len(nkey) > best_len:
                best = orig
                best_len = len(nkey)
    if best:
        return best

    # 2) fuzzy จาก pattern ที่เหมือน point_id
    cand = re.findall(r"[A-Z]{1,3}_[A-Z0-9]{1,10}(?:_[A-Z0-9]{1,10}){1,5}", t)
    if not cand:
        return None

    best_score = 0.0
    best_pid = None
    for c in cand[:12]:
        for nkey, orig in norm_map.items():
            sc = SequenceMatcher(None, c, nkey).ratio()
            if sc > best_score:
                best_score = sc
                best_pid = orig

    return best_pid if best_score >= 0.78 else None

def extract_point_id_from_image(image_bytes: bytes, norm_map: dict):
    """คืนค่า (point_id หรือ None, ocr_text ที่ใช้)"""
    # pass1: OCR เฉพาะช่วงล่างก่อน
    btm = _crop_bottom_bytes(image_bytes, frac=0.40)
    txt, _err = _vision_read_text(btm)
    pid = find_point_id_from_text(txt, norm_map)
    if pid:
        return pid, txt

    # pass2: fallback OCR ทั้งภาพ
    txt2, _err2 = _vision_read_text(image_bytes)
    pid2 = find_point_id_from_text(txt2, norm_map)
    return pid2, txt2

    
# =========================================================
# --- ✅ HISTORY GUARD (for cumulative meters) ---
# =========================================================

@st.cache_data(ttl=300)
def load_dailyreadings_tail(limit=4000):
    """โหลด DailyReadings เฉพาะท้าย ๆ เพื่อลดเวลา/โควต้า"""
    sh = gc.open(DB_SHEET_NAME)
    ws = sh.worksheet("DailyReadings")
    vals = ws.get_all_values()
    if not vals:
        return pd.DataFrame()

    header = vals[0]
    rows = vals[1:]
    if limit and len(rows) > limit:
        rows = rows[-limit:]
    df = pd.DataFrame(rows, columns=header)
    return df

def _norm_pid(pid: str) -> str:
    return str(pid or "").strip().upper()

def is_cumulative_meter(config: dict) -> bool:
    """
    Heuristic: มิเตอร์สะสม (Totalizer / เลขมิเตอร์ / kWh / total / m3 ฯลฯ)
    ปลอดภัย: ใช้กับค่าที่มักเป็น "สะสม" และ decimals == 0 เป็นหลัก
    """
    name = str(config.get("name", "") or "")
    typ  = str(config.get("type", "") or "")
    kw   = str(config.get("keyword", "") or "")
    blob = f"{name} {typ} {kw}".lower()

    # เน้นมิเตอร์สะสมมากกว่า flowrate/pressure
    hit = any(k in blob for k in [
        "เลขมิเตอร์", "meter", "total", "totalizer", "tot", "kwh", "m3", "m³"
    ])
    dec = int(config.get("decimals", 0) or 0)

    # ถ้าเป็นดิจิทัลทศนิยม (pressure/flowrate) มักไม่ใช่สะสม
    if dec > 0:
        return False

    return hit

def get_last_good_value(point_id: str, upto_date):
    """
    คืนค่า Manual_Value ล่าสุด (ไม่ใช่ FLAGGED) ที่ timestamp <= upto_date 23:59:59
    """
    df = load_dailyreadings_tail(limit=4000)
    if df.empty:
        return None

    pid = _norm_pid(point_id)

    # normalize columns
    if "point_id" not in df.columns or "timestamp" not in df.columns:
        return None

    df["point_id"] = df["point_id"].astype(str).map(_norm_pid)
    df["Status"] = df.get("Status", "").astype(str).str.strip().str.upper()

    # filter pid
    df = df[df["point_id"] == pid]
    if df.empty:
        return None

    # drop flagged
    df = df[~df["Status"].str.contains("FLAGGED", na=False)]
    if df.empty:
        return None

    # parse timestamp
    df["timestamp_dt"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp_dt"])
    if df.empty:
        return None

    cutoff = pd.to_datetime(str(upto_date) + " 23:59:59")
    df = df[df["timestamp_dt"] <= cutoff]
    if df.empty:
        return None

    df = df.sort_values("timestamp_dt")
    last = pd.to_numeric(df.iloc[-1].get("Manual_Value", None), errors="coerce")
    if pd.isna(last):
        return None
    return float(last)

def estimate_max_delta(point_id: str, upto_date, fallback=20000, max_cap=500000):
    """
    ✅ ปรับปรุง: ประเมินเพดานการเพิ่มต่อวันจากประวัติ
    
    คำนวณเพดาน 3 วิธี:
    1) Q95 (95th percentile): ทำให้ยืดหยุ่น แต่ไม่ให้ outlier ทำลาย
    2) Median * 6: ยอมค่า spike บ้าง
    3) Max daily increase * 1.2: ยอมค่าสูงสุดที่เหนือกว่า spike ได้เล็กน้อย
    
    เอาค่าที่ใหญ่ที่สุดจากสามวิธี ไว้ให้ยืดหยุ่น
    """
    df = load_dailyreadings_tail(limit=8000)
    if df.empty:
        return fallback

    pid = _norm_pid(point_id)
    if "point_id" not in df.columns or "timestamp" not in df.columns:
        return fallback

    df["point_id"] = df["point_id"].astype(str).map(_norm_pid)
    df["Status"] = df.get("Status", "").astype(str).str.strip().str.upper()
    df = df[df["point_id"] == pid]
    df = df[~df["Status"].str.contains("FLAGGED", na=False)]
    if df.empty:
        return fallback

    df["timestamp_dt"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df = df.dropna(subset=["timestamp_dt"])
    cutoff = pd.to_datetime(str(upto_date) + " 23:59:59")
    df = df[df["timestamp_dt"] <= cutoff]
    if df.empty:
        return fallback

    df = df.sort_values("timestamp_dt").tail(60)  # เอาล่าสุด 60 รายการ (ประมาณ 2 เดือน)
    vals = pd.to_numeric(df.get("Manual_Value", None), errors="coerce").dropna().astype(float).tolist()
    if len(vals) < 4:
        return fallback

    diffs = []
    for a, b in zip(vals[:-1], vals[1:]):
        d = b - a
        if d >= 0:  # เอาเฉพาะค่าที่เพิ่มขึ้น (มิเตอร์สะสม)
            diffs.append(d)

    if len(diffs) < 3:
        return fallback

    diffs = np.array(diffs, dtype=float)
    
    # ✅ ใหม่: คำนวณเพดาน 3 วิธี
    q95 = float(np.quantile(diffs, 0.95))
    med = float(np.median(diffs))
    max_diff = float(np.max(diffs))
    
    # วิธี 1: Q95 * 3 (ปกติ)
    est1 = q95 * 3.0
    
    # วิธี 2: Median * 6 (ยอมค่า spike)
    est2 = med * 6.0
    
    # วิธี 3: Max * 1.2 (ยอมค่าสูงสุดขึ้นอีกนิด)
    est3 = max_diff * 1.2
    
    # เอาค่าใหญ่ที่สุด (ให้ยืดหยุ่น)
    est = max(est1, est2, est3)
    
    # ปลอดภัย: อย่างน้อย fallback * 0.25, อย่างมาก max_cap
    est = min(est, float(max_cap))
    est = max(est, float(fallback * 0.25))
    
    return int(est)

def pick_by_history(best_val: float, candidates: list, prev_val: float, max_delta: int):
    """
    ✅ ปรับปรุง: เลือก candidate โดยพิจารณา:
      1) Original AI score (candidate["score"])
      2) History compatibility score (ความสอดคล้องกับแนวโน้มเมื่อวาน)
      3) ตรวจสอบ growth rate ให้สมเหตุสมผล
    
    คืนค่า: (picked_value, message, changed_from_ai)
    """
    if prev_val is None or not candidates:
        return best_val, "", False

    lo = float(prev_val)
    hi = float(prev_val) + float(max_delta)

    # ✅ ใหม่: คำนวณ history compatibility score สำหรับทุก candidate
    scored_cands = []
    for c in candidates:
        try:
            v = float(c.get("val"))
        except Exception:
            continue

        ai_score = float(c.get("score", 0))
        
        # History compatibility: ว่าค่า v สอดคล้องกับแนวโน้มเมื่อวานแค่ไหน
        # - ถ้า v อยู่ใน [prev_val, prev_val + max_delta] -> compatible
        # - ถ้า v ต่ำกว่า prev_val -> อาจผิด (มิเตอร์สะสมไม่ลดลง)
        # - ถ้า v สูงกว่า prev_val + max_delta -> อาจ outlier หรือกระโดด
        
        if v < lo:
            # ❌ ต่ำกว่าเมื่อวาน: ลดคะแนนมากๆ (มิเตอร์สะสมไม่ควรลดลง)
            hist_score = -1000.0
        elif v <= hi:
            # ✅ อยู่ในช่วง [prev_val, prev_val + max_delta]: ให้คะแนนเต็ม
            # - ยิ่งใกล้ prev_val มากเท่าไหร่ -> ให้คะแนนสูงขึ้น (ระดับปกติ)
            # - ยิ่งใกล้ hi มากเท่าไหร่ -> ให้คะแนนต่ำกว่า (เพิ่มมากผิดปกติ)
            delta_from_prev = v - lo
            ratio_in_range = delta_from_prev / float(max_delta)  # 0.0 = เพิ่มน้อยสุด, 1.0 = เพิ่มสูงสุด
            
            # หากอยู่ใกล้เดิม (ratio < 0.3) -> ให้คะแนนสูง (ปกติ)
            # หากอยู่ตรงกลาง (0.3 <= ratio < 0.7) -> ให้คะแนนปานกลาง
            # หากอยู่เกือบสูงสุด (ratio >= 0.7) -> ให้คะแนนต่ำ (เพิ่มมากผิดปกติ)
            if ratio_in_range < 0.3:
                hist_score = 300.0 + (0.3 - ratio_in_range) * 500.0  # 300-450
            elif ratio_in_range < 0.7:
                hist_score = 200.0 - (ratio_in_range - 0.3) * 200.0  # 200-80
            else:
                hist_score = 80.0 - (ratio_in_range - 0.7) * 400.0   # 80-0
        else:
            # ⚠️ สูงกว่าเพดาน: ลดคะแนนบ้าง แต่ไม่ถึงปฏิเสธ (อาจเป็นของจริง)
            excess = v - hi
            hist_score = max(0.0, 150.0 - excess * 2.0)
        
        # รวม AI score + history score (AI score 70% + history score 30%)
        combined_score = (ai_score * 0.7) + (hist_score * 0.3)
        scored_cands.append({
            **c,
            "val": v,
            "ai_score": ai_score,
            "hist_score": hist_score,
            "combined_score": combined_score,
        })

    if not scored_cands:
        return best_val, "", False

    # เลือก candidate ที่มี combined_score สูงสุด
    scored_cands.sort(key=lambda x: x["combined_score"], reverse=True)
    picked_c = scored_cands[0]
    picked_val = float(picked_c["val"])
    hist_sc = float(picked_c["hist_score"])
    comb_sc = float(picked_c["combined_score"])

    changed = (best_val is not None) and (abs(picked_val - float(best_val)) > 1e-9)
    
    # ✅ ข้อความ feedback อัดเพิ่มข้อมูล
    if picked_val >= lo and picked_val <= hi:
        delta_pct = ((picked_val - lo) / float(max_delta) * 100) if max_delta > 0 else 0
        msg = f"✅ History Guard: เลือกเลข {picked_val:.0f} (เพิ่มจากเมื่อวาน {delta_pct:.0f}% | combined_score={comb_sc:.0f})"
    elif picked_val < lo:
        msg = f"⚠️ History Guard: เลือก {picked_val:.0f} (ต่ำกว่าเมื่อวาน {lo:.0f}) - อาจผิด ให้ตรวจสอบ!"
        return best_val, msg, False  # ปฏิเสธ candidate นี้
    else:
        msg = f"⚠️ History Guard: เลือก {picked_val:.0f} (สูงกว่าเพดาน {hi:.0f}) - เพิ่มมากผิดปกติ"
    
    return picked_val, msg, changed

def apply_history_guard(point_id: str, best_val: float, candidates: list, config: dict, selected_date):
    """
    ✅ ปรับปรุง: ใช้ History Guard กับมิเตอร์สะสม
    
    คืนค่า: (final_value, message)
    - message: ข้อมูล feedback ที่อัดเต็มไป (ระดับการเปลี่ยนแปลง, ความมั่นใจ ฯลฯ)
    """
    if not is_cumulative_meter(config):
        return best_val, ""

    prev = get_last_good_value(point_id, selected_date - timedelta(days=1))
    if prev is None:
        return best_val, "ℹ️ ไม่มีข้อมูลประวัติ (วันแรกจึงใช้ค่า AI ตรง ๆ)"

    max_delta = estimate_max_delta(point_id, selected_date - timedelta(days=1), fallback=20000)
    picked, msg, _changed = pick_by_history(best_val, candidates, prev_val=prev, max_delta=max_delta)
    
    # ✅ เพิ่ม feedback ว่า "ตรวจสอบจำนวนเมื่อวาน"
    if msg.startswith("✅"):
        msg += f" (เมื่อวาน={prev:.0f})"
    
    return picked, msg

# =========================================================
# --- UI LOGIC ---
# =========================================================
def reset_emp_meter_state():
    st.session_state.emp_ai_value = None
    st.session_state.emp_img_hash = ""
    st.session_state.emp_ai_msg = ""
    
    # ✅ สำคัญ: เพิ่ม nonce เพื่อบังคับ Streamlit สร้าง widget ใหม่
    st.session_state.emp_nonce = int(st.session_state.get("emp_nonce", 0)) + 1

mode = st.sidebar.radio(
    "🔧 เลือกโหมดการทำงาน",
    ["📝 พนักงานจดมิเตอร์",
     "📸 อัปโหลดรูปทั้งวัน (มี point_id ในรูป)",
     "📥 อัปโหลด Excel (SCADA Export)",
     "🖥️ Dashboard Screenshot (OCR)",
     "�️ SQL Server (CUTEST SCADA - Test)",
     "�👮‍♂️ Admin Approval"]
)
if mode == "📝 พนักงานจดมิเตอร์":
    st.title("Smart Meter System")
    st.markdown("### Water treatment Plant - Borthongindustrial")
    st.caption("Version 6.2 (QR-first for Mobile + Skip Confirm)")

    # --- session state ---
    if 'confirm_mode' not in st.session_state: st.session_state.confirm_mode = False
    if 'warning_msg' not in st.session_state: st.session_state.warning_msg = ""
    if 'last_manual_val' not in st.session_state: st.session_state.last_manual_val = 0.0

    if "emp_step" not in st.session_state: st.session_state.emp_step = "SCAN_QR"
    if "emp_point_id" not in st.session_state: st.session_state.emp_point_id = ""

    # ✅ Step A: เพิ่ม nonce state (วางตรงนี้)
    if "emp_nonce" not in st.session_state:
        st.session_state.emp_nonce = 0

    all_meters = load_points_master()
    if not all_meters:
        st.error("❌ โหลด PointsMaster ไม่ได้")
        st.stop()

    # --- ฟอร์มบนสุด (มือถือควรให้สั้น) ---
    c_insp, c_date = st.columns(2)
    with c_insp:
        inspector = st.text_input("ชื่อผู้ตรวจ", "Admin", key="emp_inspector")
    with c_date:
        selected_date = st.date_input(
            "📅 วันที่จดบันทึก (ลงย้อนหลังได้)",
            value=get_thai_time().date(),
            key="emp_date"
        )

    # =========================================================
    # ✅ (2) Progress + Missing Alert (Sidebar)
    # =========================================================
    prog = get_waterreport_progress_snapshot(selected_date)
    done_set = set(prog.get("done_set") or [])
    done_val_map = dict(prog.get("value_map") or {})
    total = int(prog.get("total", 0) or 0)
    filled = int(prog.get("filled", 0) or 0)
    ratio = (filled / total) if total else 0.0
    st.sidebar.caption(
        f"ตั้ง report_col แล้ว: {int(prog.get('total_report',0) or 0)} | "
        f"ยังไม่ตั้ง: {int(prog.get('config_missing',0) or 0)}"
    )

    st.sidebar.markdown("## ✅ ความคืบหน้าการลงค่า (วันนี้)")
    st.sidebar.progress(ratio)
    st.sidebar.write(f"ลงแล้ว **{filled}/{total} จุด** ({ratio*100:.1f}%)")

    if prog.get("ok"):
        st.sidebar.caption(f"Sheet: {prog.get('sheet_title')} | Row: {prog.get('row')} | อัปเดต: {prog.get('asof')}")
    else:
        st.sidebar.error("อ่านความคืบหน้าไม่สำเร็จ")
        st.sidebar.caption(str(prog.get("error", ""))[:300])

    missing_list = prog.get("missing") or []
    if missing_list:
        with st.sidebar.expander(f"🚨 ยังไม่ลง ({len(missing_list)}) จุด", expanded=False):
            show_n = 40
            for m in missing_list[:show_n]:
                nm = m.get("name") or ""
                st.write(f"- {m['point_id']}" + (f" — {nm}" if nm else ""))
            if len(missing_list) > show_n:
                st.caption(f"...อีก {len(missing_list)-show_n} จุด")

        # Quick jump ไปจุดที่ยังไม่ลง (ช่วยให้ทีมทำงานเร็วขึ้น)
        miss_ids = [m["point_id"] for m in missing_list if m.get("point_id")]
        jump_pid = st.sidebar.selectbox("ไปยังจุดที่ยังไม่ลง", options=["(เลือก)"] + miss_ids, key="emp_jump_missing")
        if st.sidebar.button("➡️ ไปจุดนี้", use_container_width=True, key="emp_jump_btn"):
            if jump_pid != "(เลือก)":
                reset_emp_meter_state()
                st.session_state.emp_point_id = str(jump_pid).strip().upper()
                st.session_state.emp_step = "INPUT"
                st.session_state.confirm_mode = False
                st.rerun()
    else:
        st.sidebar.success("ครบแล้ว 🎉")
 
    # ถ้าอยู่โหมด mismatch confirm ให้ล็อกอยู่จุดเดิม
    if st.session_state.get("confirm_mode", False):
        st.session_state.emp_point_id = st.session_state.get("last_point_id", st.session_state.emp_point_id)
        st.session_state.emp_step = "INPUT"

    # =========================================================
    # STEP 1: SCAN QR
    # =========================================================
    if st.session_state.emp_step == "SCAN_QR":
        st.subheader("ขั้นที่ 1: สแกน QR ที่มิเตอร์")
        st.write("📌 ถ่ายให้ใกล้ ๆ และชัด (ประมาณ 15–25 ซม.)")

        tabs = st.tabs(["📷 ถ่ายด้วยกล้อง", "🖼️ อัปโหลดรูป QR (สำหรับทำบนคอม)"])
        with tabs[0]:
            qr_pic = st.camera_input("ถ่าย QR ให้ชัด", key=f"emp_qr_cam_{st.session_state.emp_nonce}")
        with tabs[1]:
            qr_upload = st.file_uploader(
                "อัปโหลดรูป QR (jpg/png)",
                type=["jpg", "jpeg", "png"],
                key=f"emp_qr_upload_{st.session_state.emp_nonce}",
                help="เหมาะสำหรับกรณีทำงานบนคอม/ไม่มี camera_input"
            )

        qr_bytes = None
        if qr_pic is not None:
            qr_bytes = qr_pic.getvalue()
        elif qr_upload is not None:
            qr_bytes = qr_upload.getvalue()

        if qr_bytes:
            pid = decode_qr(qr_bytes)
            if pid:
                reset_emp_meter_state()
                st.session_state.emp_point_id = pid
                st.session_state.emp_step = "INPUT"
                st.rerun()
            else:
                st.warning("ยังอ่าน QR ไม่ได้: ลองถ่าย/อัปโหลดใหม่ให้ชัดขึ้น หรือครอปให้เห็นเฉพาะ QR")

        # --- ทางหนีฉุกเฉิน (ซ่อน) ---
        with st.expander("สแกนไม่ได้? พิมพ์รหัสเอง"):
            manual_pid = st.text_input("พิมพ์ point_id", key="emp_manual_pid")
            if st.button("ยืนยันรหัส", use_container_width=True, key="emp_manual_ok"):
                if manual_pid.strip():
                    reset_emp_meter_state()
                    st.session_state.emp_point_id = manual_pid.strip().upper()
                    st.session_state.emp_step = "INPUT"
                    st.rerun()
                else:
                    st.warning("กรุณาพิมพ์รหัสก่อน")

        st.stop()

    # =========================================================
    # STEP 2: CONFIRM POINT (show name + ref image)
    # =========================================================
    if st.session_state.emp_step == "CONFIRM_POINT":
        pid = st.session_state.emp_point_id
        config = get_meter_config(pid)
        if not config:
            st.error(f"❌ ไม่พบ point_id: {pid} ใน PointsMaster")
            if st.button("กลับไปสแกนใหม่", use_container_width=True):
                st.session_state.emp_step = "SCAN_QR"
                st.session_state.emp_point_id = ""
                st.rerun()
            st.stop()

        meter_type = infer_meter_type(config)

        st.subheader("ขั้นที่ 2: ยืนยันจุดตรวจ")
        st.write(f"**Point:** {pid}")
        if config.get("name"):
            st.write(f"**ชื่อจุด:** {config.get('name')}")
        st.write(f"**ประเภท:** {'💧 Water' if meter_type=='Water' else '⚡ Electric'}")
        # รูปตัวอย่าง (โหลดจาก GCS ด้วยสิทธิ์ service account ไม่ต้อง public)
        ref_bytes, ref_path = load_ref_image_bytes_any(pid)
        if ref_bytes:
            st.image(ref_bytes, caption=f"รูปตัวอย่าง (Reference): {ref_path}", use_container_width=True)
        else:
            st.warning("⚠️ ไม่พบรูปตัวอย่างใน bucket สำหรับจุดนี้")
            st.caption("แนะนำให้มีไฟล์ขึ้นต้นด้วย point_id เช่น CH_S11D_106_....jpg หรือทำไฟล์มาตรฐาน ref_images/CH_S11D_106.jpg")


        b1, b2 = st.columns(2)
        if b1.button("✅ ใช่จุดนี้", type="primary", use_container_width=True):
            st.session_state.emp_step = "INPUT"
            st.rerun()
        if b2.button("❌ ไม่ใช่ / สแกนใหม่", use_container_width=True):
            st.session_state.emp_step = "SCAN_QR"
            st.session_state.emp_point_id = ""
            st.rerun()

        st.stop()

    # =========================================================
    # STEP 3: INPUT + PHOTO + SAVE
    # =========================================================
    # มาถึงตรงนี้ = emp_step == "INPUT"
    point_id = st.session_state.emp_point_id
    config = get_meter_config(point_id)
    if not config:
        st.error("❌ ไม่พบ config ของจุดนี้")
        st.session_state.emp_step = "SCAN_QR"
        st.session_state.emp_point_id = ""
        st.stop()

    report_col = str(config.get('report_col', '-') or '-').strip()
    meter_type = infer_meter_type(config)

    # =========================================================
    # ✅ (3) Duplicate Guard: เตือนถ้าวันนี้ลงค่าใน WaterReport แล้ว
    # =========================================================
    pid_u = str(point_id).strip().upper()
    if pid_u in done_set:
        existing_val = done_val_map.get(pid_u, "")
        st.warning(
            f"⚠️ จุดนี้มีค่าถูกลงใน WaterReport ของวันที่ {selected_date.strftime('%Y-%m-%d')} แล้ว"
            + (f" (ค่าในช่องตอนนี้: {existing_val})" if str(existing_val).strip() else "")
            + "\n\nถ้าจะลงซ้ำ กรุณาตรวจสอบก่อน (หรือไปทำจุดที่ยังไม่ลงจาก sidebar)"
        )
        
    st.write("---")
    c1, c2 = st.columns([2, 1])

    with c1:
        st.markdown(f"📍 จุดตรวจ: **{point_id}**")
        if config.get("name"):
            st.caption(config.get("name"))
        st.markdown(f"💾 บันทึกลงคอลัมน์: <span class='report-badge'>{report_col}</span>", unsafe_allow_html=True)
        if st.button("🔁 เปลี่ยนจุด (สแกนใหม่)", use_container_width=True, key="emp_change_point"):
            reset_emp_meter_state()
            st.session_state.emp_step = "SCAN_QR"
            st.session_state.emp_point_id = ""
            st.session_state.confirm_mode = False
            st.rerun()

    with c2:
        decimals = int(config.get("decimals", 0) or 0)
        step = 1.0 if decimals == 0 else (10 ** (-decimals))
        # ✅ Dynamic format: รองรับ decimals ทุกจำนวน (0, 1, 2, 3, ...)
        # ✅ Use % style for Streamlit st.number_input
        fmt = f"%.{decimals}f"
        st.caption("ถ่ายรูปแล้ว AI จะเสนอค่าด้านล่าง")

    # --- (Optional) แสดงรูปตัวอย่างของจุดนี้ เพื่อช่วยเช็คว่าถ่ายถูกมิเตอร์ ---
    with st.expander("🖼️ ดูรูปตัวอย่างของจุดนี้ (ถ้าต้องการเช็ค)"):
        ref_bytes, ref_path = load_ref_image_bytes_any(point_id)
        if ref_bytes:
            st.image(ref_bytes, caption="Reference: " + str(ref_path), use_container_width=True)
        else:
            st.info("ยังไม่มีรูปตัวอย่าง (Reference) ของจุดนี้ใน bucket")

    tab_cam, tab_up = st.tabs(["📷 ถ่ายรูป", "📂 อัปโหลด"])

    with tab_cam:
        img_cam = st.camera_input("ถ่ายภาพมิเตอร์", key=f"emp_meter_cam_{st.session_state.emp_nonce}")

    with tab_up:
        img_up = st.file_uploader(
            "เลือกรูปภาพ",
            type=['jpg', 'png', 'jpeg'],
            key=f"emp_meter_upload_{st.session_state.emp_nonce}"
        )
        if img_up is not None:
            st.image(img_up, caption=f"รูปที่เลือก: {getattr(img_up, 'name', 'upload')}", use_container_width=True)

    img_file = img_cam if img_cam is not None else img_up

    st.write("---")
    st.subheader("ขั้นที่ 2: ถ่ายภาพ/อัปโหลด → AI เสนอค่า → บันทึก")

    # --- กัน OCR รันซ้ำเวลาหน้า rerun ---
    if "emp_ai_value" not in st.session_state:
        st.session_state.emp_ai_value = None
    if "emp_img_hash" not in st.session_state:
        st.session_state.emp_img_hash = ""
    if "emp_ai_msg" not in st.session_state:
         st.session_state.emp_ai_msg = ""

    if img_file is None:
        st.info("📷 ถ่ายรูป (หรืออัปโหลดรูป) แล้ว AI จะอ่านค่าให้เองอัตโนมัติ")
        st.stop()

    img_bytes = img_file.getvalue()
    img_hash = hashlib.md5(img_bytes).hexdigest()

    # ถ้ารูปเปลี่ยน → อ่านใหม่
    if img_hash != st.session_state.emp_img_hash:
        st.session_state.emp_img_hash = img_hash
        with st.spinner("🤖 AI กำลังอ่านค่า..."):
            best, cand = ocr_process(img_bytes, config, debug=False, return_candidates=True)
            
            # ✅ ใช้ History Guard เฉพาะมิเตอร์สะสม (ไม่กระทบจุดอื่น)
            best2, msg = best, ""
            try:
                best2, msg = apply_history_guard(point_id, best, cand, config, selected_date)
            except Exception:
                best2, msg = best, ""
                
            st.session_state.emp_ai_value = float(best2)
            st.session_state.emp_ai_msg = msg

    # --- FIX: กันค่า AI ติดลบไม่ให้ st.number_input ล้ม ---
    ai_val = float(st.session_state.emp_ai_value or 0.0)

    min_allowed = 0.0
    prefill_val = ai_val if ai_val >= min_allowed else min_allowed
    if ai_val < min_allowed:
        st.warning("⚠️ AI อ่านค่าได้ติดลบ (น่าจะอ่านผิด) — ระบบจะให้แก้เองก่อนบันทึก")

    st.write(f"🤖 **AI เสนอค่า:** {fmt % ai_val}")
    if st.session_state.get("emp_ai_msg"):
        st.info(st.session_state.emp_ai_msg)

    choice = st.radio(
        "จะบันทึกค่าไหน?",
        ["✅ ใช้ค่า AI", "✍️ แก้เอง"],
        horizontal=True,
        key="emp_choice"
    )

    if choice == "✍️ แก้เอง":
        final_val = st.number_input(
            "พิมพ์ค่าที่ถูกต้อง",
            value=float(prefill_val),
            min_value=min_allowed,
            step=step,
            format=fmt,
            key="emp_override_val"
        )
        status = "CONFIRMED_MANUAL"
    else:
        if ai_val < min_allowed:
            st.error("❌ AI อ่านค่าได้ติดลบ จึงไม่อนุญาตให้บันทึกแบบ 'ใช้ค่า AI' — กรุณาเลือก '✍️ แก้เอง'")
            st.stop()
        final_val = float(ai_val)
        status = "CONFIRMED_AI"

    st.info(f"ค่าที่จะบันทึก: {fmt % float(final_val)}")

    col_save, col_retry = st.columns(2)

    if col_save.button("💾 บันทึกค่า", type="primary", use_container_width=True):
        try:
            filename = f"{point_id}_{selected_date.strftime('%Y%m%d')}_{get_thai_time().strftime('%H%M%S')}.jpg"
            image_url = upload_image_to_storage(img_bytes, filename)

            ok = save_to_db(point_id, inspector, meter_type, float(final_val), float(ai_val), status, selected_date, image_url)
            if ok:
                ok_r, msg_r = export_to_real_report(point_id, float(final_val), inspector, report_col, selected_date, debug=True)
                if not ok_r:
                    st.warning('⚠️ ส่งค่าไป TEST waterreport ไม่สำเร็จ: ' + msg_r)
                st.success("✅ บันทึกสำเร็จ")

                # ไปจุดถัดไป
                reset_emp_meter_state()
                st.session_state.emp_step = "SCAN_QR"
                st.session_state.emp_point_id = ""
                st.rerun()
            else:
                st.error("❌ Save Failed")
        except Exception as e:
            st.error(f"❌ บันทึกไม่สำเร็จ: {e}")

    if col_retry.button("🔁 ถ่าย/เลือกใหม่", use_container_width=True):
        reset_emp_meter_state()
        st.rerun()

elif mode == "📸 อัปโหลดรูปทั้งวัน (มี point_id ในรูป)":
    st.title("📸 อัปโหลดรูปทั้งวัน → อ่าน point_id → ลง WaterReport")
    st.caption("อัปโหลดรูปหลายไฟล์/zip ที่มี point_id อยู่ในรูป แล้วระบบจะอ่านค่าให้และบันทึกลง WaterReport แบบครั้งเดียว")

    c_insp, c_date = st.columns(2)
    with c_insp:
        inspector = st.text_input("ชื่อผู้บันทึก", "Admin", key="bulk_inspector")
    with c_date:
        # ✅ ไม่ใช้ key ให้ date picker อ่านค่า default ทุกครั้ง (ไม่ cache)
        report_date = st.date_input("📅 วันที่ของรายงาน", value=get_thai_time().date())

    norm_map = build_pid_norm_map()
    pm = load_points_master() or []
    all_pids = sorted({str(r.get("point_id","")).strip().upper() for r in pm if r.get("point_id")})

    up_files = st.file_uploader(
        "อัปโหลดรูป (หลายไฟล์) หรือ zip",
        type=["jpg","jpeg","png","zip"],
        accept_multiple_files=True,
        key="bulk_upload"
    )
    if not up_files:
        st.stop()

    # ✅ ตรวจสอบว่ากลุ่มไฟล์เปลี่ยนหรือเปล่า → ถ้าเปลี่ยน reset bulk_rows
    current_file_ids = tuple(sorted([f.name for f in up_files]))
    if "bulk_last_files" not in st.session_state:
        st.session_state["bulk_last_files"] = None
    
    if st.session_state["bulk_last_files"] != current_file_ids:
        # ✅ ไฟล์เปลี่ยนแล้ว → reset all state
        st.session_state["bulk_rows"] = None
        st.session_state["bulk_expanded"] = {}
        st.session_state["bulk_last_files"] = current_file_ids

    # แตกไฟล์: รองรับ zip + รูปตรง ๆ
    images = []  # [{name, bytes}]
    for f in up_files:
        name = getattr(f, "name", "upload")
        b = f.getvalue()
        if name.lower().endswith(".zip"):
            z = zipfile.ZipFile(io.BytesIO(b))
            for zi in z.infolist():
                if zi.filename.lower().endswith((".jpg",".jpeg",".png")):
                    images.append({"name": os.path.basename(zi.filename), "bytes": z.read(zi)})
        else:
            images.append({"name": name, "bytes": b})

    st.write(f"พบรูปทั้งหมด: **{len(images)}** ไฟล์")
    st.session_state["bulk_image_map"] = {it["name"]: it["bytes"] for it in images}

    if "bulk_rows" not in st.session_state:
        st.session_state["bulk_rows"] = None

    # ✅ อัตโนมัติ process ทันทีที่อัพเสร็จ (ไม่ต้องกดปุ่ม)
    if st.session_state["bulk_rows"] is None:
        st.info("🔄 กำลังประมวลผลรูป...")
        rows = []
        
        # สร้าง progress bar ด้านนอก loop เพื่อให้เห็น realtime
        progress_container = st.empty()
        status_container = st.empty()
        
        for i, it in enumerate(images, start=1):
            img_name = it["name"]
            img_bytes = it["bytes"]

            # Update progress text
            status_container.text(f"📍 กำลังประมวลผล: {i}/{len(images)} - {img_name[:40]}")
            progress_container.progress(i / len(images))

            pid, _pid_text = extract_point_id_from_image(img_bytes, norm_map)
            pid_u = str(pid).strip().upper() if pid else ""

            cfg = get_meter_config(pid_u) if pid_u else None
            ai_val = None
            msg = ""
            stt = "NO_PID"
            candidates_list = []

            if pid_u and cfg:
                try:
                    best, cand = ocr_process(img_bytes, cfg, return_candidates=True)
                    # ✅ ใช้ History Guard + candidates
                    best2, hmsg = apply_history_guard(pid_u, best, cand, cfg, report_date)
                    ai_val = float(best2)
                    msg = hmsg or ""
                    stt = "OK"
                    # เก็บ candidates เผื่อผู้ใช้ต้องการดูทางเลือก
                    candidates_list = cand if isinstance(cand, list) else []
                except Exception as e:
                    stt = "OCR_FAIL"
                    msg = str(e)[:200]
            elif pid_u and not cfg:
                stt = "NO_CONFIG"
                msg = "ไม่พบ config ของจุดนี้ใน PointsMaster"

            rows.append({
                "file": img_name,
                "point_id": pid_u or "",
                "ai_value": ai_val,
                "final_value": ai_val,
                "status": stt,
                "note": msg,
                "candidates": candidates_list,
                "image_bytes": img_bytes,  # ✅ เก็บรูปไว้
            })

        progress_container.empty()
        status_container.empty()
        
        st.session_state["bulk_rows"] = rows
        st.session_state["bulk_candidates_storage"] = {rows[i]["file"]: rows[i].get("candidates", []) for i in range(len(rows))}
        st.success(f"✅ ประมวลผลเสร็จ {len(rows)} รูป")
        st.rerun()

    rows = st.session_state.get("bulk_rows")
    if not rows:
        st.stop()

    st.subheader("📊 ผลการประมวลผลและแก้ไข")
    
    # ✅ สรุปผลการอ่าน
    ok_count = sum(1 for r in rows if r.get("status") == "OK")
    no_pid = sum(1 for r in rows if r.get("status") == "NO_PID")
    no_cfg = sum(1 for r in rows if r.get("status") == "NO_CONFIG")
    fail = sum(1 for r in rows if r.get("status") == "OCR_FAIL")
    
    col_ok, col_pid, col_cfg, col_fail = st.columns(4)
    col_ok.metric("✅ OK", ok_count)
    col_pid.metric("🚫 ไม่มี ID", no_pid)
    col_cfg.metric("⚙️ ไม่มี config", no_cfg)
    col_fail.metric("❌ Fail", fail)
    
    st.divider()
    
    # ✅ ตัวเก็บสำหรับจำ expand state
    if "bulk_expanded" not in st.session_state:
        st.session_state["bulk_expanded"] = {}
    
    # ✅ แสดงตาราง compact พร้อมปุ่ม expand ที่ละแถว
    st.markdown("### 🖼️ แต่ละรูป (คลิก 'ดูรายละเอียด' เพื่อแก้ไข)")
    
    for idx, r in enumerate(rows):
        file_name = r.get("file", "")
        pid = r.get("point_id", "")
        val = r.get("final_value", 0)
        status = r.get("status", "")
        note = r.get("note", "")
        img_bytes = r.get("image_bytes")
        candidates = r.get("candidates", [])
        
        # Status emoji
        status_emoji = {
            "OK": "✅",
            "SAVED": "💾",
            "NO_PID": "🚫",
            "NO_CONFIG": "⚙️",
            "OCR_FAIL": "❌",
        }.get(status, "❓")
        
        # Compact display
        col1, col2, col3, col4, col5 = st.columns([2, 2, 1.5, 1.5, 1], gap="small")
        
        with col1:
            st.caption(f"📄 {file_name[:30]}")
        with col2:
            st.caption(f"**{pid}**" if pid else "—")
        with col3:
            # ✅ แสดงค่า 0 ได้ (เช่นค่าเริ่มต้น) + format ตามจำนวนทศนิยม
            if val is not None and str(val).strip() != "":
                cfg = get_meter_config(pid)
                decimals = int(cfg.get('decimals', 0) or 0) if cfg else 0
                fmt = f"{{:.{decimals}f}}"
                st.caption(f"ค่า: **{fmt.format(val)}**")
            else:
                st.caption("—")
        with col4:
            st.caption(f"{status_emoji} {status}")
        with col5:
            expand_key = f"expand_{idx}"
            if st.button("📋", key=f"btn_{idx}", help="ดูรายละเอียด"):
                st.session_state["bulk_expanded"][expand_key] = not st.session_state["bulk_expanded"].get(expand_key, False)
                st.rerun()
        
        # ✅ Expandable detail view
        if st.session_state["bulk_expanded"].get(expand_key, False):
            with st.container(border=True):
                det_col1, det_col2 = st.columns([2, 1], vertical_alignment="top")
                
                # แสดงรูป
                with det_col1:
                    st.markdown("#### 🖼️ รูปต้นฉบับ")
                    if img_bytes:
                        st.image(img_bytes, use_container_width=True)
                    else:
                        st.warning("ไม่มีรูป")
                
                # แสดง candidates + แก้ค่า + บันทึก
                with det_col2:
                    # ✅ ถ้าเป็น NO_PID ให้พิมพ์ point_id เอง
                    if status == "NO_PID":
                        st.markdown("#### 📍 พิมพ์ point_id")
                        manual_pid_input = st.selectbox(
                            "เลือก point_id จากรายชื่อ",
                            options=["(พิมพ์เอง)"] + all_pids,
                            key=f"no_pid_select_{idx}"
                        )
                        
                        if manual_pid_input == "(พิมพ์เอง)":
                            manual_pid_text = st.text_input(
                                "พิมพ์ point_id",
                                value="",
                                placeholder="เช่น GU_BP_3_2",
                                key=f"no_pid_input_{idx}"
                            )
                            selected_manual_pid = manual_pid_text.strip().upper() if manual_pid_text else ""
                        else:
                            selected_manual_pid = manual_pid_input
                        
                        if st.button("✅ ยืนยัน point_id", key=f"confirm_pid_{idx}", use_container_width=True, type="primary"):
                            if selected_manual_pid:
                                # เช็ค config
                                cfg = get_meter_config(selected_manual_pid)
                                if cfg:
                                    rows[idx]["point_id"] = selected_manual_pid
                                    rows[idx]["status"] = "OK"
                                    rows[idx]["note"] = f"✅ Manual input: {selected_manual_pid}"
                                    st.session_state["bulk_rows"] = rows
                                    st.success(f"✅ เปลี่ยนเป็น {selected_manual_pid}")
                                    st.rerun()
                                else:
                                    st.error(f"❌ ไม่พบ config สำหรับ {selected_manual_pid}")
                            else:
                                st.warning("กรุณาพิมพ์ point_id")
                        
                        st.divider()
                    
                    st.markdown("#### 📋 ทางเลือกค่า")
                    
                    # Show candidates top 3
                    if candidates:
                        st.caption("Top candidates:")
                        try:
                            for c_idx, c in enumerate(candidates[:3]):
                                c_val = float(c.get("val", 0))
                                c_score = float(c.get("score", 0))
                                # ✅ Format ตามจำนวนทศนิยม
                                cfg = get_meter_config(rows[idx].get("point_id", ""))
                                decimals = int(cfg.get('decimals', 0) or 0) if cfg else 0
                                fmt = f"{{:.{decimals}f}}"
                                val_str = fmt.format(c_val)
                                if st.button(f"ใช้ {val_str} (score {c_score:.0f})", key=f"use_cand_{idx}_{c_idx}", use_container_width=True):
                                    rows[idx]["final_value"] = c_val
                                    st.session_state["bulk_rows"] = rows
                                    st.success(f"✅ เปลี่ยนเป็น {val_str}")
                                    st.rerun()
                        except Exception as e:
                            st.error(f"❌ ข้อผิดพลาด: {e}")
                    else:
                        st.info("ไม่มี candidates")
                    
                    st.divider()
                    
                    # Manual edit
                    st.caption("📝 พิมพ์เอง:")
                    
                    # ✅ Get point_id first (selectbox comes first)
                    new_pid = st.selectbox(
                        "point_id",
                        options=[""] + all_pids,
                        index=([""] + all_pids).index(str(rows[idx].get("point_id", "")).strip().upper()) 
                               if str(rows[idx].get("point_id", "")).strip().upper() in ([""] + all_pids) else 0,
                        key=f"manual_pid_{idx}"
                    )
                    
                    # ✅ Get decimals to format correctly (now new_pid is defined)
                    try:
                        cfg_manual = get_meter_config(new_pid or rows[idx].get("point_id", ""))
                        decimals_manual = int(cfg_manual.get('decimals', 0) or 0) if cfg_manual else 0
                        step_manual = 1.0 if decimals_manual == 0 else (10 ** (-decimals_manual))
                        # ✅ Use % style format for Streamlit st.number_input
                        fmt_manual = f"%.{decimals_manual}f"
                        
                        # ✅ Safe get final_value with default 0
                        try:
                            final_value_current = float(rows[idx].get("final_value") or 0)
                        except (ValueError, TypeError):
                            final_value_current = 0.0
                        
                        new_val = st.number_input(
                            "ค่าใหม่",
                            value=final_value_current,
                            step=step_manual,
                            format=fmt_manual,
                            key=f"manual_val_{idx}"
                        )
                    except Exception as e:
                        st.error(f"❌ เกิดข้อผิดพลาดในการแสดงช่องกรอกค่า: {e}")
                        st.stop()
                    
                    if st.button("💾 บันทึกแก้ไข", key=f"save_{idx}", use_container_width=True, type="primary"):
                        rows[idx]["final_value"] = new_val
                        rows[idx]["point_id"] = new_pid
                        st.session_state["bulk_rows"] = rows
                        st.success("✅ บันทึกแก้ไขแล้ว")
                        st.session_state["bulk_expanded"][expand_key] = False
                        st.rerun()
                    
                    st.divider()
                    
                    # ✅ ปุ่มบันทึกลง Google Sheet ทันที
                    st.markdown("#### 📊 บันทึกลง Google Sheet")
                    
                    final_pid = new_pid or rows[idx].get("point_id", "")
                    final_val = new_val if new_val is not None else rows[idx].get("final_value", None)
                    
                    # ✅ แก้ validation: เช็ค None/empty string แทน 0
                    if not final_pid:
                        st.warning("ต้องมีค่า point_id ก่อน")
                    elif final_val is None or str(final_val).strip() == "":
                        st.warning("ต้องมีค่าตัวเลขก่อน")
                    else:
                        col_save, col_skip = st.columns(2)
                        
                        with col_save:
                            if st.button("✅ บันทึกทันที", key=f"save_sheet_{idx}", type="primary", use_container_width=True):
                                if not final_pid or final_val is None or str(final_val).strip() == "":
                                    st.error("❌ ต้องมีค่า + point_id ก่อน")
                                else:
                                    cfg = get_meter_config(final_pid)
                                    if not cfg:
                                        st.error("❌ ไม่พบ config")
                                    else:
                                        report_col = str(cfg.get("report_col", "") or "").strip()
                                        if not report_col or report_col in ("-", "—", "–"):
                                            st.error("❌ ไม่มี report_col")
                                        else:
                                            # ✅ Parse value to float
                                            try:
                                                write_val = float(str(final_val).replace(",", "").strip())
                                            except Exception:
                                                write_val = str(final_val).strip()
                                            
                                            # ✅ Use inspector from parent scope
                                            inspector_name = str(inspector).strip() or "Admin"
                                            
                                            # บันทึกลง Google Sheet
                                            ok_r, msg_r = export_to_real_report(
                                                final_pid, 
                                                write_val, 
                                                inspector_name, 
                                                report_col, 
                                                report_date, 
                                                debug=True
                                            )
                                        
                                        if ok_r:
                                            st.success(f"✅ บันทึกสำเร็จ: {msg_r}")
                                            rows[idx]["status"] = "SAVED"  # Mark as saved
                                            st.session_state["bulk_rows"] = rows
                                            st.session_state["bulk_expanded"][expand_key] = False
                                            st.rerun()
                                        else:
                                            st.error(f"❌ บันทึกไม่สำเร็จ: {msg_r}")
                        
                        with col_skip:
                            if st.button("⏭️ ข้าม", key=f"skip_{idx}", use_container_width=True):
                                st.session_state["bulk_expanded"][expand_key] = False
                                st.rerun()
                
                st.caption(f"📌 Status: {status}")
                if note:
                    st.caption(f"💬 {note}")
        
        st.divider()

    write_mode_ui = st.radio(
        "เวลาบันทึกให้ทำแบบไหน?",
        ["เขียนทับทั้งหมด", "เขียนเฉพาะช่องว่าง (ไม่ทับของเดิม)"],
        index=0,
        horizontal=True,
        key="bulk_write_mode",
    )

    # ✅ ปุ่มบันทึกที่เหลือ (ที่ยังไม่บันทึก)
    rows_final = st.session_state.get("bulk_rows", rows)
    unsaved_count = sum(1 for r in rows_final if r.get("status") != "SAVED")
    
    if unsaved_count > 0:
        if st.button(f"✅ บันทึกที่เหลือ ({unsaved_count} จุด)", type="primary", use_container_width=True):
            report_items = []
            db_rows = []
            fail_list = []

            folder = f"daily_bulk/{report_date.strftime('%Y%m%d')}"
            inspector_name = inspector or "Admin"

            for r in rows_final:
                if r.get("status") == "SAVED":
                    continue  # ข้ามที่บันทึกแล้ว
                
                pid_u = str(r.get("point_id","")).strip().upper()
                val = r.get("final_value", None)

                if not pid_u or val is None or str(val).strip() == "":
                    continue

                cfg = get_meter_config(pid_u)
                if not cfg:
                    fail_list.append((pid_u, "NO_CONFIG_IN_PointMaster"))
                    continue

                report_col = str(cfg.get("report_col","")).strip()
                if not report_col or report_col in ("-","—","–"):
                    fail_list.append((pid_u, "NO_REPORT_COL"))
                    continue

                img_bytes = r.get("image_bytes")

                image_url = "-"
                if img_bytes:
                    pid_slug = pid_u.replace(" ", "_")
                    filename = f"{folder}/{pid_slug}_{get_thai_time().strftime('%H%M%S')}.jpg"
                    image_url = upload_image_to_storage(img_bytes, filename)

                try:
                    write_val = float(str(val).replace(",", "").strip())
                except Exception:
                    write_val = str(val).strip()

                report_items.append({"point_id": pid_u, "value": write_val, "report_col": report_col})

                try:
                    meter_type = infer_meter_type(cfg)
                except Exception:
                    meter_type = "Electric"

                record_ts = datetime.combine(report_date, get_thai_time().time()).strftime("%Y-%m-%d %H:%M:%S")
                db_rows.append([record_ts, meter_type, pid_u, inspector_name, write_val, write_val, "AUTO_BULK_IMAGE_OCR", image_url])

            if not report_items:
                st.info("ไม่มีข้อมูลที่เหลือให้บันทึก")
            else:
                ok_db, db_msg = append_rows_dailyreadings_batch(db_rows)
                if not ok_db:
                    st.warning(f"⚠️ Log DailyReadings ไม่สำเร็จ: {db_msg}")

                wm = "overwrite" if write_mode_ui.startswith("เขียนทับ") else "empty_only"
                ok_pids, fail_report = export_many_to_real_report_batch(report_items, report_date, debug=True, write_mode=wm)

                # ✅ 1.3: Update logging with results
                if HAS_LOGGER:
                    update_log_success(ok_pids)
                    if fail_report:
                        update_log_failed(fail_report)
                    # Show daily report
                    st.info(print_daily_report())

                # ✅ Show results
                if ok_pids:
                    st.success(f"✅ ลง WaterReport สำเร็จ: {len(ok_pids)} จุด")
                
                if fail_report:
                    st.error(f"❌ ไม่สำเร็จ: {len(fail_report)} จุด")
                    with st.expander("📋 ดูรายละเอียด:"):
                        for pid, reason in fail_report:
                            st.caption(f"  • {pid}: {reason}")
                
                # ✅ Update rows status carefully
                for r in rows_final:
                    pid_u = str(r.get("point_id","")).strip().upper()
                    if pid_u in ok_pids:
                        r["status"] = "SAVED"
                    elif any(pid_u == str(f[0]).strip().upper() for f in fail_report):
                        r["status"] = "ERROR"
                
                st.session_state["bulk_rows"] = rows_final
                st.rerun()
    else:
        st.success("✅ ทั้งหมดบันทึกแล้ว!")

elif mode == "🖥️ Dashboard Screenshot (OCR)":
    st.title("🖥️ Dashboard Screenshot → WaterReport")
    st.caption("อัปโหลดรูปหน้าจอ Dashboard แล้วระบบจะอ่านค่า Pressure/Flowrate/Flow_Total ของ FLOW 1-3")

    c_insp, c_date = st.columns(2)
    with c_insp:
        inspector = st.text_input("ชื่อผู้บันทึก", "Admin", key="dash_inspector")
    with c_date:
        report_date = st.date_input("📅 วันที่ของรายงาน (ที่จะไปกรอกใน WaterReport)", value=get_thai_time().date(), key="dash_date")

    up = st.file_uploader("อัปโหลดรูปหน้าจอ Dashboard (JPG/PNG)", type=["jpg", "jpeg", "png"], key="dash_img")
    if not up:
        st.info("อัปโหลดรูปก่อน แล้วกดปุ่มอ่านค่า")
        st.stop()

    img_bytes = up.getvalue()
    st.image(img_bytes, caption=f"ภาพที่อัปโหลด: {getattr(up, 'name', 'dashboard')}", use_container_width=True)

    # กัน OCR รันซ้ำตอน rerun
    if "dash_img_hash" not in st.session_state:
        st.session_state.dash_img_hash = ""
    if "dash_rows" not in st.session_state:
        st.session_state.dash_rows = None
    if "dash_dbg" not in st.session_state:
        st.session_state.dash_dbg = None

    img_hash = hashlib.md5(img_bytes).hexdigest()
    if img_hash != st.session_state.dash_img_hash:
        st.session_state.dash_img_hash = img_hash
        st.session_state.dash_rows = None
        st.session_state.dash_dbg = None

    if st.button("🔎 อ่านค่าจากรูป (OCR)"):
        with st.spinner("กำลังอ่านค่าจากรูป..."):
            rows, dbg = extract_dashboard_flow_values(img_bytes, debug=True)
        st.session_state.dash_rows = rows
        st.session_state.dash_dbg = dbg

    rows = st.session_state.dash_rows
    if not rows:
        st.stop()

    st.subheader("ผลการอ่านค่าจากรูป")
    st.dataframe(pd.DataFrame(rows), use_container_width=True)

    pm = load_points_master()
    all_pids = sorted({str(r.get("point_id", "")).strip().upper() for r in pm if r.get("point_id")})

    st.subheader("แปลงค่าที่อ่านได้ → point_id")
    picked = []

    for r in rows:
        flow_label = r.get("flow", "")
        try:
            n = int(str(flow_label).strip().split()[-1])
        except Exception:
            n = None

        st.markdown(f"#### {flow_label}")
        cols = st.columns(3)

        metrics = [
            ("pressure_bar", "Pressure (bar)"),
            ("flowrate_m3h", "Flowrate (m3/h)"),
            ("flow_total_m3", "Flow_Total (m3)"),
        ]

        for i, (k, label) in enumerate(metrics):
            v = r.get(k)
            with cols[i]:
                st.caption(label)
                st.write(v)

                default_pid = (_DASH_DEFAULT_POINT_MAP.get((n, k), "") if n else "")
                default_pid = str(default_pid).strip().upper()

                options = ["(ไม่บันทึก)"] + all_pids
                default_idx = options.index(default_pid) if default_pid in options else 0

                sel = st.selectbox(
                    "point_id",
                    options=options,
                    index=default_idx,
                    key=f"dash_pid_{flow_label}_{k}"
                )

                if sel != "(ไม่บันทึก)" and v is not None:
                    picked.append({"point_id": sel, "value": v})

    with st.expander("Debug OCR"):
        st.json(st.session_state.dash_dbg or {})

    st.subheader("บันทึกลง WaterReport")
    if st.button("✅ บันทึกลง WaterReport (อัตโนมัติ)"):
        inspector_name = inspector or "Admin"

        report_items = []
        db_rows = []
        fail_list = []

        for it in picked:
            pid_u = str(it.get("point_id", "")).strip().upper()
            val = it.get("value", None)
            if not pid_u or val is None or str(val).strip() == "":
                continue

            cfg = get_meter_config(pid_u)
            if not cfg:
                fail_list.append((pid_u, "NO_CONFIG_IN_POINTSMaster"))
                continue

            report_col = str(cfg.get("report_col", "") or "").strip()
            if (not report_col) or (report_col in ("-", "—", "–")):
                fail_list.append((pid_u, "NO_REPORT_COL_IN_POINTSMaster"))
                continue

            try:
                write_val = float(str(val).replace(",", "").strip())
            except Exception:
                write_val = str(val).strip()

            report_items.append({"point_id": pid_u, "value": write_val, "report_col": report_col})

            try:
                meter_type = infer_meter_type(cfg)
            except Exception:
                meter_type = "Electric"

            current_time = get_thai_time().time()
            record_ts = datetime.combine(report_date, current_time).strftime("%Y-%m-%d %H:%M:%S")
            db_rows.append([record_ts, meter_type, pid_u, inspector_name, write_val, write_val, "AUTO_DASHBOARD_OCR", "-"])

        if not report_items:
            st.warning("ไม่มีข้อมูลให้บันทึก")
            st.stop()

        ok_db, db_msg = append_rows_dailyreadings_batch(db_rows)
        if not ok_db:
            st.warning(f"⚠️ Log ลง DailyReadings ไม่สำเร็จ: {db_msg}")

        with st.spinner("กำลังบันทึกลง WaterReport..."):
            ok_pids, fail_report = export_many_to_real_report_batch(report_items, report_date, debug=True)

        # ✅ 1.3: Update logging with results
        if HAS_LOGGER:
            update_log_success(ok_pids)
            if fail_report:
                update_log_failed(fail_report)
            # Show daily report
            st.info(print_daily_report())

        st.success(f"✅ บันทึกสำเร็จ: {len(ok_pids)} จุด")
        if fail_report:
            st.error(f"❌ ไม่สำเร็จ: {len(fail_report)} จุด")
            st.write([[pid, reason] for pid, reason in fail_report])

elif mode == "�️ SQL Server (CUTEST SCADA - Test)":
    st.title("🗄️ SQL Server Integration (Test Mode)")
    st.markdown("### ดึงข้อมูลจาก CUTEST SCADA 2018 SQL Server โดยตรง")
    
    st.warning("⚠️ นี่คือโหมดทดสอบ (Test Mode) - ยังไม่เก็บข้อมูลลง Google Sheet")
    
    # ตรวจสอบว่ามี driver ที่ติดตั้ง
    has_driver = HAS_PYMSSQL or HAS_PYODBC or HAS_SQLALCHEMY
    if not has_driver:
        st.error("""
        ❌ ต้องติดตั้ง SQL Driver อย่างน้อยตัวหนึ่ง:
        
        **แนะนำสำหรับ macOS/Linux:**
        ```
        pip install pymssql
        ```
        
        **หรือ (ทั่วไป):**
        ```
        pip install sqlalchemy
        ```
        
        **หรือ (Windows):**
        ```
        pip install pyodbc
        ```
        """)
        st.stop()
    
    # แสดง driver ที่ใช้อยู่
    driver_list = []
    if HAS_PYMSSQL: driver_list.append("✅ pymssql")
    if HAS_SQLALCHEMY: driver_list.append("✅ sqlalchemy")
    if HAS_PYODBC: driver_list.append("✅ pyodbc")
    st.info(f"🔌 Driver ที่พบ: {', '.join(driver_list)}")
    
    st.markdown("---")
    st.subheader("1️⃣ ตั้งค่าเชื่อมต่อ SQL Server")
    
    col1, col2 = st.columns(2)
    with col1:
        sql_server = st.text_input("Server Address", placeholder="192.168.1.100 หรือ localhost", value=st.session_state.get("sql_server", ""))
        sql_username = st.text_input("Username", placeholder="sa", value=st.session_state.get("sql_username", ""))
    
    with col2:
        sql_database = st.text_input("Database Name", placeholder="CUTEST_DB", value=st.session_state.get("sql_database", ""))
        sql_password = st.text_input("Password", type="password", placeholder="password", value=st.session_state.get("sql_password", ""))
    
    if st.button("💾 บันทึกการตั้งค่า"):
        st.session_state["sql_server"] = sql_server
        st.session_state["sql_database"] = sql_database
        st.session_state["sql_username"] = sql_username
        st.success("✅ บันทึกการตั้งค่าแล้ว")
    
    if st.button("🔌 ทดสอบเชื่อมต่อ", type="primary"):
        if not sql_server or not sql_database or not sql_username:
            st.error("❌ กรุณากรอกข้อมูลให้ครบ")
        else:
            with st.spinner("กำลังทดสอบเชื่อมต่อ..."):
                success, message = test_sql_connection(sql_server, sql_database, sql_username, sql_password)
                if success:
                    st.success(message)
                else:
                    st.error(message)
    
    st.markdown("---")
    st.subheader("2️⃣ ดึงค่าจาก SCADA")
    
    all_meters = load_points_master()
    meter_choices = {m['point_id']: f"{m['point_id']} - {m.get('device_name', '')}" for m in all_meters if m.get('point_id')}
    
    if not meter_choices:
        st.warning("❌ ไม่พบ Point ID ใน PointsMaster")
    else:
        selected_pid = st.selectbox(
            "เลือก Point ID",
            options=list(meter_choices.keys()),
            format_func=lambda x: meter_choices[x]
        )
        
        col_date, col_time = st.columns(2)
        with col_date:
            query_date = st.date_input("เลือกวันที่", value=get_thai_time().date())
        with col_time:
            query_time = st.text_input("เวลา (HH:MM) - ไม่บังคับ", placeholder="14:30")
        
        if st.button("🔍 ดึงข้อมูล", type="primary"):
            if not sql_server or not sql_database or not sql_username:
                st.error("❌ กรุณากรอกข้อมูล SQL Server ให้ครบ")
            else:
                with st.spinner("กำลังดึงข้อมูล..."):
                    result = query_scada_values(
                        sql_server,
                        sql_database,
                        sql_username,
                        sql_password,
                        selected_pid,
                        query_date,
                        query_time
                    )
                
                if result["success"]:
                    st.success("✅ ดึงข้อมูลสำเร็จ!")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("ค่าล่าสุด", f"{result.get('value', 'N/A')}")
                    with col2:
                        st.metric("เวลา", result.get('timestamp', 'N/A')[:19])
                    with col3:
                        st.metric("จำนวนแถว", result.get('record_count', 0))
                    
                    st.info(f"📊 จากตาราง: {result.get('table', 'Unknown')}")
                    
                    if result.get('all_records'):
                        st.subheader("📋 ข้อมูลทั้งหมด")
                        df_records = pd.DataFrame(result['all_records'])
                        st.dataframe(df_records, use_container_width=True)
                        
                        csv_data = df_records.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            "📥 ดาวน์โหลด CSV",
                            csv_data,
                            f"scada_{selected_pid}_{query_date}.csv",
                            "text/csv"
                        )
                else:
                    st.error(result.get('message', "เกิดข้อผิดพลาด"))
    
    st.markdown("---")
    st.subheader("📝 หมายเหตุ")
    st.info("""
    **CUTEST SCADA 2018 ใช้ตาราง SQL:**
    - `History_Data` - ข้อมูลประวัติหลัก
    - `Readings` - ข้อมูลการอ่าน
    - `TagHistory` - ประวัติ Tag
    
    **คอลัมน์ที่ต้องมี:**
    - `TagName` หรือ `PointID` - รหัสจุดวัด
    - `Value` - ค่าที่วัด
    - `Timestamp` - เวลา
    """)

elif mode == "�👮‍♂️ Admin Approval":
    st.title("👮‍♂️ Admin Dashboard")
    st.caption("ระบบอนุมัติผลการอ่านค่ามิเตอร์น้ำ/ไฟ")
    if st.button("🔄 รีเฟรช"): st.rerun()

    sh = gc.open(DB_SHEET_NAME)
    ws = sh.worksheet("DailyReadings")
    data = ws.get_all_records()
    pending = [d for d in data if str(d.get('Status', '')).strip().upper() == 'FLAGGED']

    if not pending: st.success("✅ All Clear")
    else:
        for i, item in enumerate(pending):
            with st.container():
                st.markdown("---")
                c_info, c_val, c_act = st.columns([1.5, 1.5, 1])
                with c_info:
                    st.subheader(f"🚩 {item.get('point_id')}")
                    st.caption(f"Inspector: {item.get('inspector')}")
                    img_url = item.get('image_url')
                    if img_url and img_url != '-' and str(img_url).startswith('http'):
                        st.image(img_url, width=220)
                    else:
                        st.warning("No Image")

                with c_val:
                    m_val = safe_float(item.get('Manual_Value'), 0.0)
                    a_val = safe_float(item.get('AI_Value'), 0.0)
                    options_map = {
                        f"👤 คนจด: {m_val}": m_val,
                        f"🤖 AI: {a_val}": a_val
                    }
                    selected_label = st.radio("เลือกค่าที่ถูกต้อง:", list(options_map.keys()), key=f"rad_{i}")
                    choice = options_map[selected_label]

                with c_act:
                    st.write("")
                    if st.button("✅ อนุมัติ", key=f"btn_{i}", type="primary"):
                        try:
                            timestamp = str(item.get('timestamp', '')).strip()
                            point_id = str(item.get('point_id', '')).strip()
                            cells = ws.findall(timestamp)
                            updated = False
                            for cell in cells:
                                if str(ws.cell(cell.row, 3).value).strip() == point_id:
                                    ws.update_cell(cell.row, 7, "APPROVED")
                                    ws.update_cell(cell.row, 5, choice)
                                    config = get_meter_config(point_id)
                                    report_col = (config.get('report_col', '') if config else '')
                                    
                                    # ✅ Parse timestamp กลับเป็นวันที่ เพื่อหา Sheet ให้เจอตอน Approve
                                    try:
                                        dt_obj = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                                        approve_date = dt_obj.date()
                                    except:
                                        approve_date = get_thai_time().date() # fallback
                                        
                                    ok_r, msg_r = export_to_real_report(point_id, choice, str(item.get('inspector', '')), report_col, approve_date, debug=True)
                                    if not ok_r:
                                        st.warning('⚠️ ส่งค่าไป TEST waterreport ไม่สำเร็จ: ' + msg_r)
                                    updated = True; break
                            if updated: st.success("Approved!"); st.rerun()
                            else: st.warning("หา row ไม่เจอ")
                        except Exception as e: st.error(f"Error approve: {e}")

elif mode == "📥 อัปโหลด Excel (SCADA Export)":
    st.title("📥 อัปโหลด Excel (SCADA Export)")
    st.caption("โหมดนี้ใช้แทนการถ่ายรูป SCADA: เอาไฟล์ Excel ที่ SCADA export มาอัปโหลด แล้วระบบจะดึงค่า + บันทึกลง WaterReport ให้อัตโนมัติ")

    st.info(
        "วิธีใช้ (แบบง่าย ป.6)\n"
        "1) กด 'Browse files' แล้วเลือกไฟล์ Excel ที่ลูกค้าส่งมา (เลือกได้หลายไฟล์)\n"
        "2) เลือก 'วันที่ของรายงาน' ให้ตรงกับวันที่ใน WaterReport\n"
        "3) กดปุ่ม 'ดึงค่าจาก Excel' เพื่อให้ระบบอ่านค่า\n"
        "4) ถ้ามีจุดที่ไม่มีใน Excel -> กรอกเองเฉพาะจุดนั้น\n"
        "5) กด 'บันทึกลง WaterReport' จบ ✅"
    )

    # เลือกวันที่รายงาน
    report_date = st.date_input("📅 วันที่ของรายงาน (ที่จะไปกรอกใน WaterReport)", value=get_thai_time().date())
    report_date_str = report_date.strftime("%Y/%m/%d")

    # โหลด mapping
    st.subheader("1) ไฟล์ Mapping (DB_Water_Scada.xlsx)")
    mapping_rows = []
    if os.path.exists("DB_Water_Scada.xlsx"):
        st.success("พบไฟล์ DB_Water_Scada.xlsx ในโปรเจกต์ ✅ (จะใช้ไฟล์นี้อัตโนมัติ)")
        mapping_rows = load_scada_excel_mapping(local_path="DB_Water_Scada.xlsx")
    else:
        st.warning("ไม่พบไฟล์ DB_Water_Scada.xlsx ในโปรเจกต์ — กรุณาอัปโหลดไฟล์นี้ก่อน (ไฟล์เล็ก ๆ)")
        uploaded_map = st.file_uploader("อัปโหลด DB_Water_Scada.xlsx", type=["xlsx"])
        if uploaded_map is not None:
            mapping_rows = load_scada_excel_mapping(uploaded_bytes=uploaded_map.getvalue())

    if not mapping_rows:
        st.stop()

    # อัปโหลด Excel export (จำไฟล์ได้ / เพิ่มไฟล์ทีหลังได้)
    st.subheader("2) อัปโหลดไฟล์ Excel ที่ SCADA export (เพิ่มทีหลังได้)")

    import hashlib, time

    # --- โหลดย้อนหลังไฟล์ที่เคยอัปโหลด (ในรอบนี้) ---
    if "scada_files" not in st.session_state:
        # filename -> {bytes, sha1, size, added_at, processed_sha1}
        st.session_state["scada_files"] = {}
    if "excel_updated_pids_last_run" not in st.session_state:
        st.session_state["excel_updated_pids_last_run"] = []

    # 2.1 อัปโหลดไฟล์ใหม่ (จะถูก 'เพิ่ม' เข้า list เดิม ไม่ทับ)
    exports_new = st.file_uploader(
        "เลือกไฟล์ Excel (เลือกได้หลายไฟล์) เช่น ...Daily_Report.xlsx, ...UF_System.xlsx, ...SMMT_Daily_Report.xlsx",
        type=["xlsx"],
        accept_multiple_files=True,
        key="scada_exports_uploader",
    )

    added_count = 0
    if exports_new:
        for f in exports_new:
            b = f.getvalue()
            h = hashlib.sha1(b).hexdigest()
            old = st.session_state["scada_files"].get(f.name)
            if (old is None) or (old.get("sha1") != h):
                st.session_state["scada_files"][f.name] = {
                    "bytes": b,
                    "sha1": h,
                    "size": len(b),
                    "added_at": time.time(),
                    "processed_sha1": (old or {}).get("processed_sha1"),
                }
                added_count += 1

    if added_count:
        st.success(f"เพิ่มไฟล์ใหม่ {added_count} ไฟล์ ✅ (ไฟล์เดิมยังอยู่)")

    files_dict = st.session_state.get("scada_files", {})
    if not files_dict:
        st.info("ยังไม่มีไฟล์ Excel — กรุณาอัปโหลดอย่างน้อย 1 ไฟล์")
        st.stop()

    # 2.2 แสดงรายการไฟล์ที่มีอยู่
    def _is_new_file(meta: dict) -> bool:
        return (meta or {}).get("processed_sha1") != (meta or {}).get("sha1")

    file_rows = []
    for name, meta in files_dict.items():
        file_rows.append({
            "ไฟล์": name,
            "ขนาด(MB)": round((meta.get("size", 0) or 0) / 1_000_000, 2),
            "สถานะ": "NEW" if _is_new_file(meta) else "มีแล้ว",
        })

    st.dataframe(pd.DataFrame(file_rows), use_container_width=True)

    c1, c2, c3 = st.columns([2, 1, 1])
    with c1:
        remove_sel = st.multiselect("เลือกลบไฟล์ (ถ้าต้องการ)", options=list(files_dict.keys()), default=[])
    with c2:
        if st.button("🗑️ ลบไฟล์ที่เลือก"):
            for fn in remove_sel:
                files_dict.pop(fn, None)
            st.session_state["scada_files"] = files_dict
            st.rerun()
    with c3:
        if st.button("🧹 ล้างไฟล์ทั้งหมด"):
            st.session_state["scada_files"] = {}
            st.session_state.pop("excel_results", None)
            st.session_state.pop("excel_missing", None)
            st.session_state["excel_updated_pids_last_run"] = []
            st.rerun()

    # 2.3 เลือกว่าจะให้ระบบอ่านไฟล์แบบไหน
    st.markdown("### 2.3 เลือกโหมดการอ่านไฟล์")
    process_mode = st.radio(
        "จะให้ระบบอ่านไฟล์แบบไหน?",
        ["📚 อ่านทุกไฟล์ที่มี", "➕ อ่านเฉพาะไฟล์ใหม่ (NEW)", "🎯 อ่านเฉพาะไฟล์ที่เลือก"],
        index=0,
        horizontal=True,
        key="scada_process_mode",
    )

    # ⚙️ ตั้งค่า max_scan_rows
    with st.expander("⚙️ ตั้งค่าการสแกน (กรณีดึงค่าไม่ครบ):"):
        scan_option = st.radio(
            "จำนวนแถวที่สแกนต่อไฟล์?",
            ["🚀 สแกนแบบเร็ว (5,000 แถว)", "⚖️ สแกนตัวกลาง (50,000 แถว)", "🔍 สแกนเต็มที่ (100,000+ แถว)", "⚡ สแกนโดยไม่จำกัด (เนื่องจากอาจช้า)"],
            index=1,  # ค่าเริ่มต้น: ตัวกลาง
            horizontal=False,
        )
        
        max_scan_rows_custom = 0  # ค่าเริ่มต้น = ปล่อยให้ระบบเลือก
        if "🚀" in scan_option:
            max_scan_rows_custom = 5000
        elif "⚖️" in scan_option:
            max_scan_rows_custom = 50000
        elif "🔍" in scan_option:
            max_scan_rows_custom = 500000  # เพิ่มจาก 100000 เป็น 500000
        elif "⚡" in scan_option:
            max_scan_rows_custom = 999999999  # ไม่จำกัด
            st.warning("⚠️ โหมด 'ไม่จำกัด' อาจทำให้ระบบช้า หากไฟล์มีแถวมากมาย (เช่น > 500,000 แถว)")
        
        st.caption(f"ค่าปัจจุบัน: {max_scan_rows_custom:,} แถว" if max_scan_rows_custom else "ค่าปัจจุบัน: ปล่อยให้ระบบเลือก")

    all_files = list(files_dict.keys())
    new_files = [fn for fn in all_files if _is_new_file(files_dict.get(fn, {}))]

    proc_files = []
    if process_mode.startswith("📚"):
        proc_files = all_files
    elif process_mode.startswith("➕"):
        proc_files = new_files
        if not proc_files:
            st.warning("ยังไม่มีไฟล์ NEW ให้ประมวลผล (เพิ่มไฟล์ใหม่ก่อน หรือเลือกโหมดอื่น)")
    else:
        proc_files = st.multiselect(
            "เลือกไฟล์ที่จะให้ระบบอ่าน",
            options=all_files,
            default=all_files,
            key="scada_selected_files",
        )

    # สร้าง dict เฉพาะไฟล์ที่ต้องอ่าน (เพื่อให้ 'เพิ่มไฟล์ใหม่ทีหลัง' ไม่ต้องประมวลผลไฟล์เก่า)
    uploaded_exports_proc = {fn: files_dict[fn]["bytes"] for fn in proc_files if fn in files_dict}

    # ปุ่มดึงค่า
    if st.button("🔎 ดึงค่าจาก Excel"):
        if not uploaded_exports_proc:
            st.warning("ยังไม่ได้เลือกไฟล์ให้ประมวลผล")
            st.stop()

        with st.spinner("กำลังอ่านค่าใน Excel..."):
            # === (Optional) จับคู่ไฟล์กรณีลูกค้าเปลี่ยนชื่อ ===
            # ปกติระบบจะเดาจากชื่อไฟล์เอง แต่ถ้าขึ้น NO_FILE ให้ตั้งค่าตรงนี้
            file_key_map = {}
            key_norms = sorted({_strip_date_prefix(r.get("file_key", "")) for r in mapping_rows if r.get("file_key")})

            with st.expander("⚙️ ตั้งค่าจับคู่ไฟล์ (ใช้เมื่อขึ้น NO_FILE / ลูกค้าเปลี่ยนชื่อไฟล์)"):
                if not key_norms:
                    st.info("ไม่พบ file_key ใน mapping")
                else:
                    # **สำคัญ**: ให้เลือกได้เฉพาะไฟล์ที่กำลังประมวลผลในรอบนี้
                    options = ["(Auto)"] + list(uploaded_exports_proc.keys())

                    for kn in key_norms:
                        if not kn:
                            continue

                        default_choice = "(Auto)"
                        kn_strip = (kn or "").strip().lower()
                        kn_norm = _norm_filekey(kn_strip)

                        # 1) ตรงชื่อแบบตัดวันที่แล้ว
                        exact_cands = []
                        for fname in uploaded_exports_proc.keys():
                            f_strip = _strip_date_prefix(fname)
                            if f_strip == kn_strip:
                                exact_cands.append(fname)

                        # 2) ตรงแบบ normalize
                        if not exact_cands and kn_norm:
                            for fname in uploaded_exports_proc.keys():
                                f_strip = _strip_date_prefix(fname)
                                if _norm_filekey(f_strip) == kn_norm:
                                    exact_cands.append(fname)

                        if exact_cands:
                            if "smmt" not in kn_norm:
                                non_smmt = [f for f in exact_cands if "smmt" not in _norm_filekey(_strip_date_prefix(f))]
                                default_choice = non_smmt[0] if non_smmt else exact_cands[0]
                            else:
                                default_choice = exact_cands[0]
                        else:
                            # 3) fallback แบบ scoring
                            best = None
                            best_score = -10**9
                            for fname in uploaded_exports_proc.keys():
                                f_strip = _strip_date_prefix(fname)
                                f_norm = _norm_filekey(f_strip)
                                score = 0
                                if f_strip == kn_strip:
                                    score += 1000
                                if f_norm == kn_norm and kn_norm:
                                    score += 900
                                if kn_strip and kn_strip in f_strip:
                                    score += 80
                                if kn_norm and kn_norm in f_norm:
                                    score += 60
                                if ("smmt" in f_norm) != ("smmt" in kn_norm):
                                    score -= 500
                                if kn_norm and f_norm.startswith(kn_norm):
                                    score += 40
                                if score > best_score:
                                    best_score = score
                                    best = fname
                            if best is not None and best_score >= 200:
                                default_choice = best

                        # UF_System: ถ้าไม่มีไฟล์ UF จริง ให้เดา AF_Report/Report_Gen (ถ้ามีอยู่ในรอบนี้)
                        if default_choice == "(Auto)":
                            kn2 = _norm_filekey(kn)
                            if "uf" in kn2 or "uf_system" in kn2 or "ufsystem" in kn2:
                                for fname in uploaded_exports_proc.keys():
                                    fn2 = _norm_filekey(fname)
                                    if fn2.startswith("af_report") or "report_gen" in fn2:
                                        default_choice = fname
                                        break

                        sel = st.selectbox(
                            f"ไฟล์ที่ใช้สำหรับ '{kn}'",
                            options=options,
                            index=options.index(default_choice) if default_choice in options else 0,
                            key=f"filemap_{kn}",
                        )
                        if sel != "(Auto)":
                            file_key_map[kn] = sel

                    st.caption("ทิป: ถ้า UF/System เปลี่ยนชื่อไฟล์ ให้เลือกไฟล์ AF_Report_Gen.. มาแทนคีย์ UF_System")

            # --- ดึงค่าเฉพาะไฟล์ที่เลือก ---
            allow_single = True if process_mode.startswith("📚") else False
            results_new, missing_new = extract_scada_values_from_exports(
                mapping_rows,
                uploaded_exports_proc,
                file_key_map=file_key_map,
                target_date=report_date,
                allow_single_file_fallback=allow_single,
                custom_max_scan_rows=max_scan_rows_custom,
            )

            # --- รวมผล: ถ้าอ่านเฉพาะไฟล์ใหม่/ไฟล์ที่เลือก ให้ 'เติมเพิ่ม' โดยไม่ลบของเดิม ---
            prev = st.session_state.get("excel_results")
            merged = []
            updated_pids = set()

            if prev and (not process_mode.startswith("📚")):
                prev_by_pid = {str(r.get("point_id")): r for r in prev}
                for r in results_new:
                    pid = str(r.get("point_id"))
                    ok_new = (r.get("status") == "OK") and (r.get("value") is not None)
                    if ok_new:
                        rr = dict(r)
                        rr["_updated"] = True
                        merged.append(rr)
                        updated_pids.add(pid)
                    else:
                        old = prev_by_pid.get(pid)
                        ok_old = old and (old.get("status") == "OK") and (old.get("value") is not None)
                        if ok_old:
                            oo = dict(old)
                            oo["_updated"] = False
                            merged.append(oo)
                        else:
                            rr = dict(r)
                            rr["_updated"] = False
                            merged.append(rr)
            else:
                for r in results_new:
                    ok_new = (r.get("status") == "OK") and (r.get("value") is not None)
                    rr = dict(r)
                    rr["_updated"] = bool(ok_new)
                    merged.append(rr)
                    if ok_new:
                        updated_pids.add(str(r.get("point_id")))

            # ทำ missing จาก merged (เอาไว้ให้กรอกเอง)
            missing_point_ids = [r.get("point_id") for r in merged if not (r.get("status") == "OK" and r.get("value") is not None)]
            missing_merged = [{"point_id": pid} for pid in missing_point_ids if pid]

            # mark processed ให้ไฟล์ที่อ่านแล้ว (ไฟล์ใหม่จะกลายเป็น 'มีแล้ว')
            for fn in proc_files:
                if fn in files_dict:
                    files_dict[fn]["processed_sha1"] = files_dict[fn].get("sha1")
            st.session_state["scada_files"] = files_dict

        # เก็บไว้ใน session
        st.session_state["excel_results"] = merged
        st.session_state["excel_missing"] = missing_merged
        st.session_state["excel_updated_pids_last_run"] = sorted(list(updated_pids))

    # ถ้ามีผลแล้ว แสดงส่วนแก้/บันทึก
    if "excel_results" in st.session_state:
        results = st.session_state["excel_results"]
        missing = st.session_state.get("excel_missing", [])

        # แสดงผลสรุป + ตาราง (แสดงครั้งเดียว)
        ok_count = sum(1 for r in results if r.get("status") == "OK" and r.get("value") is not None)
        st.success(f"อ่านได้แล้ว {ok_count}/{len(results)} จุด")

        # ถ้าอ่านแบบเพิ่มไฟล์ใหม่ -> จะมีคอลัมน์ _updated เพื่อดูว่ารอบล่าสุดอัปเดตอะไรบ้าง
        show_only_updated = st.checkbox("🆕 แสดงเฉพาะจุดที่อัปเดตจากการดึงค่ารอบล่าสุด", value=False)

        show_only_missing = st.checkbox("🚫 แสดงเฉพาะจุดที่ไม่มีใน Excel", value=False)
        df_show = pd.DataFrame(results)
        if show_only_missing and (not df_show.empty) and ("status" in df_show.columns):
            df_show = df_show[df_show["status"] != "OK"]
        if show_only_updated and (not df_show.empty) and ("_updated" in df_show.columns):
            df_show = df_show[df_show["_updated"] == True]
        st.dataframe(df_show, use_container_width=True)


        # เตือนจุดที่หาย
        missing_point_ids = [m["point_id"] for m in missing]
        if missing_point_ids:
            st.warning("มีจุดที่ดึงค่าไม่สำเร็จ/ไม่มีใน Excel: " + ", ".join(missing_point_ids))

        # ให้กรอกเองเฉพาะจุดที่หาย
        manual_inputs = {}
        with st.expander("✍️ กรอกเองเฉพาะจุดที่ไม่มีใน Excel (ถ้ามี)"):
            for pid in missing_point_ids:
                manual_inputs[pid] = st.text_input(f"{pid} (กรอกตัวเลข)", value="")

        # รวมค่า final
        final_values = {}
        for r in results:
            pid = r["point_id"]
            val = r["value"]
            if pid in manual_inputs and manual_inputs[pid].strip() != "":
                val = manual_inputs[pid].strip()
            final_values[pid] = val

        # ปุ่มบันทึกลง WaterReport
        st.subheader("3) บันทึกลง WaterReport")
        st.caption("จะบันทึกเฉพาะจุดที่มีค่า (ไม่ว่าง) ลง WaterReport ตาม report_col ใน PointsMaster")

        
        # เลือกวิธีบันทึก (กันเขียนทับ / บันทึกเฉพาะรอบล่าสุด)
        save_scope = st.radio(
            "จะบันทึกข้อมูลชุดไหน?",
            ["บันทึกทุกจุดที่มีค่า", "บันทึกเฉพาะจุดที่อัปเดตจากรอบล่าสุด"],
            index=0,
            horizontal=True,
            key="scada_save_scope",
        )
        write_mode_ui = st.radio(
            "เวลาบันทึกให้ทำแบบไหน?",
            ["เขียนทับทั้งหมด", "เขียนเฉพาะช่องว่าง (ไม่ทับของเดิม)"],
            index=0,
            horizontal=True,
            key="scada_write_mode",
        )

        if st.button("✅ บันทึกลง WaterReport (อัตโนมัติ)"):
            inspector_name = "Admin"

            # 1) เตรียมรายการที่จะบันทึก (validate + แปลงค่า)
            report_items = []   # ส่งเข้า WaterReport
            db_rows = []        # log ลง DailyReadings
            fail_list = []      # [(pid, reason), ...]

            # ถ้าเลือก 'บันทึกเฉพาะจุดที่อัปเดตจากรอบล่าสุด' จะกรองจุดที่ไม่เกี่ยวออก
            last_updated = set(st.session_state.get("excel_updated_pids_last_run", []) or [])
            manual_updated = {pid for pid, vv in manual_inputs.items() if str(vv).strip() != ""}
            allowed_pids = None
            if save_scope.startswith("บันทึกเฉพาะ"):
                allowed_pids = last_updated.union(manual_updated)
                if not allowed_pids:
                    st.warning("ยังไม่มีจุดที่อัปเดตจากรอบล่าสุดให้บันทึก")
                    st.stop()

            for pid, val in final_values.items():
                pid_u = str(pid).strip().upper()
                if allowed_pids is not None and pid_u not in allowed_pids:
                    continue

                if val is None or str(val).strip() == "":
                    continue

                cfg = get_meter_config(pid_u)
                if not cfg:
                    fail_list.append((pid_u, "NO_CONFIG_IN_POINTSMaster"))
                    continue

                report_col = str(cfg.get("report_col", "") or "").strip()
                if (not report_col) or (report_col in ("-", "—", "–")):
                    fail_list.append((pid_u, "NO_REPORT_COL_IN_POINTSMaster"))
                    continue

                # แปลงค่าให้เป็นตัวเลขถ้าเป็นไปได้
                write_val = val
                try:
                    write_val = float(str(val).replace(",", "").strip())
                except Exception:
                    write_val = str(val).strip()

                report_items.append({
                    "point_id": pid_u,
                    "value": write_val,
                    "report_col": report_col
                })

                # ทำ log ลง DB (DailyReadings) — timestamp = วันที่รายงาน + เวลาปัจจุบัน (ไทย)
                try:
                    meter_type = infer_meter_type(cfg)
                except Exception:
                    meter_type = "Electric"

                try:
                    current_time = get_thai_time().time()
                    record_ts = datetime.combine(report_date, current_time).strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    record_ts = get_thai_time().strftime("%Y-%m-%d %H:%M:%S")

                db_rows.append([
                    record_ts,
                    meter_type,
                    pid_u,
                    inspector_name,
                    write_val,   # Manual_Value
                    write_val,   # AI_Value
                    "AUTO_EXCEL_SCADA",
                    "-"          # image_url
                ])

            if not report_items:
                st.warning("ไม่มีข้อมูลให้บันทึก (ค่าทั้งหมดว่าง)")
                st.stop()

            # 2) log ลง DB แบบ batch (ลด requests)
            ok_db, db_msg = append_rows_dailyreadings_batch(db_rows)
            db_ok_count = len(db_rows) if ok_db else 0
            if not ok_db:
                # ไม่หยุดระบบ แค่แจ้งให้รู้ว่าล็อก DB ไม่สำเร็จ
                st.warning(f"⚠️ Log ลง DailyReadings ไม่สำเร็จ: {db_msg}")

            # 3) export ลง WaterReport แบบ batch (ลด Read requests)
            with st.spinner("กำลังบันทึกลง WaterReport..."):
                wm = "overwrite" if write_mode_ui.startswith("เขียนทับ") else "empty_only"
                ok_pids, fail_report = export_many_to_real_report_batch(report_items, report_date, debug=True, write_mode=wm)

            # ✅ 1.3: Update logging with results
            if HAS_LOGGER:
                update_log_success(ok_pids)
                if fail_report:
                    update_log_failed(fail_report)
                # Show daily report
                st.info(print_daily_report())

            report_ok = len(ok_pids)
            report_fail = list(fail_report)

            # แยก 'ข้ามเพราะช่องมีข้อมูลแล้ว' ออกจาก error จริง
            skipped = [(pid, reason) for pid, reason in report_fail if str(reason) == 'SKIP_NON_EMPTY']
            report_fail_real = [(pid, reason) for pid, reason in report_fail if str(reason) != 'SKIP_NON_EMPTY']

            st.success(f"✅ บันทึกลง WaterReport สำเร็จ: {report_ok} จุด")
            st.info(f"🗃️ Log ลง DailyReadings สำเร็จ: {db_ok_count} จุด")

            if skipped:
                st.info(f"⏭️ ข้าม {len(skipped)} จุด เพราะช่องมีข้อมูลอยู่แล้ว (เลือกโหมด 'เขียนทับทั้งหมด' ถ้าต้องการทับ)")

            if report_fail_real:
                st.error(f"❌ บันทึกไม่สำเร็จ: {len(report_fail_real)} จุด")
                st.write([[pid, reason] for pid, reason in report_fail_real])
        st.divider()
        st.info("หมายเหตุ: ถ้าลูกค้าบอกว่า 'มีมิเตอร์ไฟ 1 จุดที่ไม่มี export มาใน Excel' -> ใช้ช่องกรอกเองด้านบนได้เลย (เหมือนมิเตอร์น้ำ)")
